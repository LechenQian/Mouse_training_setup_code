'''
Created on July 17, 2019

@author: Selina Qian
'''
from ScopeFoundry import Measurement

import numpy as np
import random


import time

from AntCamHW.daq_do.daq_do_dev import DAQSimpleDOTask
from AntCamHW.daq_di.daq_di_dev import DAQSimpleDITask

from openpyxl import Workbook

class OdorGen(object):
    def __init__(self,odorindex):
        self.odorindex = odorindex
        self.odors_DAQ = []


    def assign_odor(self):
    # initiate all the odor solenoids

        for item in self.odorindex:

            self.odors_DAQ.append(DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(item)))
        print('Odor {} has been properly assigned'.format(self.odorindex))


    def set_rewardodor(self,index):
        reward_odor = self.odors_DAQ[index]
        print('reward odor is odor {}'.format(index))
        return reward_odor


    def initiate(self):
        for odor in self.odors_DAQ:
            odor.low()
        print('Odor initiation: status low')
    def close(self):
        for odor in self.odors_DAQ:
            odor.close()
        print('Connection has been closed')


class SelinaTraining(Measurement):
    def __init__(self):
        self.events_filename = '2019-7-17-test.xlsx'
        self.reward_odor_index = 0

        self.duration_rec_off = 6.5
        self.duration_rec_on_before = 4  # change this to exponential decay
        self.duration_odor_to_outcome = 1.3
        self.duration_water_large = 0.2

        self.duration_rec_on_after = 8
        self.duration_odor_on = 0.5

    def run(self):
        print('a ha')

        odors_cue = OdorGen([0, 1, 2, 3])
        odors_cue.assign_odor()
        self.reward_odor = odors_cue.set_rewardodor(index=self.reward_odor_index)
        odors_cue.initiate()
        # odors_cue.odors_DAQ[i]
        print('odor done')

        self.waterR = DAQSimpleDOTask('Dev1/port0/line0')
        self.waterR.low()
        # self.OdorOnCopy = DAQSimpleDOTask('Dev3/port2/line5')
        # self.OdorOnCopy.low()
        self.lickR = DAQSimpleDITask('Dev2_SELECT/port1/line0')
        print('water done')

        # EVENT CODES
        # video recording start / start trial = 101
        # lick on = 11, lick off = 10

        # contingency reward odor on = 131, off = 130, water on = 51, right water off = 50
        # contingency no reward odor on = 141, off = 140, water on = 61, right water off = 60
        # non-contingency reward odor on = 151, off = 150, water on = 71, right water off = 70
        # non-contingency no reward odor on = 161, off = 160, water on = 81, right water off = 80

        # create excel workbook
        self.wb = Workbook()
        self.ws = self.wb.active
        print('book done')


        #generate trial type
        numtrials = 200
        p_cont_noncont = 0.5
        p_USwCS = 0.5
        p_USwoCS = 0.5
        cont_reward = np.zeros(int(numtrials * p_cont_noncont * p_USwCS))  # code 0
        cont_noreward = np.ones(int(numtrials * p_cont_noncont * (1 - p_USwCS)))  # code 1
        temp_comb1 = np.concatenate((cont_reward, cont_noreward))

        noncont_reward = np.ones(int(numtrials * (1 - p_cont_noncont) * p_USwoCS)) * 2  # code 2
        noncont_noreward = np.ones(int(numtrials * (1 - p_cont_noncont) * (1 - p_USwoCS))) * 3  # code 3
        temp_comb2 = np.concatenate((noncont_noreward, noncont_reward))

        trialtypes = np.concatenate((temp_comb1, temp_comb2))
        random.shuffle(trialtypes)
        print(trialtypes)

        # counters for each trial type
        self.counter = np.zeros(4)



        for t in range(0, numtrials):

            print('trial number: ', t)
            print()

            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=101)
            self.check_licking_1spout(self.duration_rec_on_before)

#           main training program
            self.run_trial_type(int(trialtypes[t]))

            self.check_licking_1spout(self.duration_rec_on_after)

            # self.settings.save_video.update_value(False):

            self.wb.save(self.events_filename)

            self.check_licking_1spout(self.duration_rec_off)



        odors_cue.initiate()
        odors_cue.close()
        self.waterR.low()
        self.waterR.close()

        print('FINISHED ASSOCIATION TRAINING')


        if self.settings.save_video.value():
            self.recorder.close()




    def check_licking_1spout(self, interval):

        checkperiod = 0.02
        timeout = time.time() + interval

        right_lick_last = 0
        while time.time() < timeout:
            right_lick = self.lickR.read()

            if right_lick != right_lick_last:
                if right_lick:

                    print('Lick')
                    d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
                    d = self.ws.cell(row=self.ws.max_row, column=2, value=11)
                else:

                    d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
                    d = self.ws.cell(row=self.ws.max_row, column=2, value=10)
            else:
                pass

            right_lick_last = right_lick
            time.sleep(checkperiod)

    def run_trial_type(self,types):
        odor_on = False
        reward_on = False
        if types == 0:
            print('contingency reward trial ' + str(int(self.counter[types])))
            print('opening odor port')
            odor_on = True
            reward_on = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_outcome)  ### can be a problem
            self.run_reward_module(reward_on, w_code)

        elif types == 1:
            print('contingency no reward trial ' + str(int(self.counter[types])))
            print('opening odor port')
            odor_on = True
            r_code = [141, 140]
            w_code = [61, 60]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_outcome)  ### can be a problem
            self.run_reward_module(reward_on, w_code)
        elif types == 2:
            print('non-contingency reward trial ' + str(int(self.counter[types])))
            reward_on = True
            r_code = [151, 150]
            w_code = [71, 70]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_outcome)  ### can be a problem
            self.run_reward_module(reward_on, w_code)
        else:
            print('non-contingency no reward trial ' + str(int(self.counter[types])))
            r_code = [161, 160]
            w_code = [71, 70]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_outcome)  ### can be a problem
            self.run_reward_module(reward_on, w_code)

        self.counter[types] += 1

        self.wb.save(self.events_filename)




    def run_odor_module(self,odor_on, r_code):
        if odor_on:
            self.reward_odor.high()
            # self.OdorOnCopy.high()  # ？？？
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[0])

            time.sleep(self.duration_odor_on)

            print('closing odor port')
            self.reward_odor.low()

            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[1])
        else:
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[0])
            time.sleep(self.duration_odor_on)
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[1])

    def run_reward_module(self,reward_on, w_code):
        if reward_on:

            # modify! give water if licks three times within 1 s

            print('opening water valve')
            self.waterR.high()
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[0])
            self.check_licking_1spout(self.duration_water_large)  # this parameter hasn't een defined

            print('closing water valve')
            self.waterR.low()
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[1])

        else:
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[0])
            time.sleep(self.duration_water_large)
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[1])

#     def camera_action(self):
#         '''
#         format the image properly
#         '''
#         try:
#             wide_image = self.wide_cam.read()
#             wide_image_data = self.wide_cam.to_numpy(wide_image)
#             self.wide_disp_queue.put(wide_image_data)
#
#             if self.settings.save_video.value() and self.settings.in_trial.value() :
#                 self.recorder.save_frame(self.settings.filename.value(),wide_image)
#
#         except Exception as ex:
#             print('Error : %s' % ex)


test = SelinaTraining()
print('start')
test.run()

