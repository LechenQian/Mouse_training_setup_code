'''
Created on July 17, 2019

@author: Selina Qian
'''
from ScopeFoundry import Measurement
import datetime
import numpy as np
import random
import pickle
import time
from AntCamHW.daq_do.daq_do_dev import DAQSimpleDOTask
from AntCamHW.daq_di.daq_di_dev import DAQSimpleDITask
from openpyxl import Workbook
import os
import logging

class OdorGen(object):
    def __init__(self,odorindex):
        self.odorindex = odorindex
        self.odors_DAQ = []


    def assign_odor(self):
        # initiate all the odor solenoids
        for item in self.odorindex:
            self.odors_DAQ.append(DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(item)))
        print('Odor {} has been properly assigned'.format(self.odorindex))

    def set_rewardodor(self, index: list):
        reward_odor = self.odors_DAQ[index[0]]
        non_reward_odor = self.odors_DAQ[index[1]]
        print('reward odor is odor {}'.format(index))
        return reward_odor,  non_reward_odor

    def initiate(self):
        for odor in self.odors_DAQ:
            odor.low()
        print('Odor initiation: status low')

    def close(self):
        for odor in self.odors_DAQ:
            odor.close()
        print('Connection has been closed')


class SelinaTraining(Measurement):
    def __init__(self):
        self.list = [7, 6]
        self.events_path = "C:/Users/MurthyLab/Desktop/Selina/experiment_data/C14/"+datetime.datetime.now().strftime("%Y-%m-%d")+"/"
        self.events_filename = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M")+'session_1.xlsx'
        self.reward_odor_index = [1, 0] #odor list index change according to mi
        self.operant = True
        self.licknum = 1


        self.numtrials = 160
        # pre training
        self.p_conbynoncon = 0.5 # total water volumn should keep the same
        self.p_reward_USwCS = 0.8 # temp a.8 from np.rand self.p_conbynoncon*self.p_reward_USwCS + (1-self.p_conbynoncon)*self.p_USwoCs = p_pretraining_go
        self.p_reward_USwoCS = 0.4
        self.counter = np.zeros(4)

        self.duration_rec_on_before = 2.5
        self.duration_odor_on = 1
        self.duration_odor_to_action = 0
        self.duration_action_window = 2.5
        self.duration_water_large = 0.1
        self.duration_rec_on_after = 3
        self.duration_ITI = np.random.exponential(2, size=self.numtrials)

        self.waterline = 0
        self.filename = self.events_path + self.events_filename
    def generate_trial(self):
        total_num = 20
        percent_1 = round((1 - self.p_conbynoncon)*self.p_reward_USwoCS,2) #non-con rewarded

        percent_3 = round((1 - self.p_conbynoncon)*(1-self.p_reward_USwoCS),2) # non-con no reward
        print(percent_3)
        percent_2by24 = round(self.p_reward_USwCS,2)
        percent_coupled_24 = percent_1
        percent_single_24 = round(1 - percent_1 - percent_coupled_24 - percent_3,2)

        percent_24 = percent_coupled_24 + percent_single_24
        print(percent_single_24)
        assert percent_single_24 > 0, print("invalid ratio")

        sum_freq = percent_1 + percent_24 + percent_3
        assert abs(sum_freq - 1) < 1e-3, print("Sum of frequent ({}) not equal to 1".format(sum_freq))

        num_1 = int(total_num * percent_1)
        num_3 = int(total_num * percent_3)
        num_4 = int(round(total_num * percent_24 * (1-percent_2by24),1))
        num_single_24 = int(total_num * percent_single_24)
        print(num_1,num_single_24,num_3,num_4)
        final_data = []
        for i in range(int(self.numtrials/total_num)):

            _data = [0] * num_1 + [2] * num_3 + [1] * num_single_24
            random.shuffle(_data)

            while _data[-1] == 0:
                random.shuffle(_data)

            data = []
            for val in _data:
                data = data + [0, 1] if val == 0 else data + [val]
            index_2 = [i for i in range(len(data)) if data[i] == 1]
            replace_index = random.sample(index_2, num_4)
            for item in replace_index:
                data[item] = 3
            print(data)
            final_data.extend(data)

        return final_data
    def run(self):
        try:
            os.mkdir(self.events_path)
        except OSError:
            print("The directory %s existed" % self.events_path)
        else:
            print("Successfully created the directory %s " % self.events_path)
        logname = self.filename[0:-5] + '.log'

        odors_cue = OdorGen(self.list)
        odors_cue.assign_odor()
        self.reward_odor, self.non_reward_odor = odors_cue.set_rewardodor(index=self.reward_odor_index)
        odors_cue.initiate()
        # odors_cue.odors_DAQ[i]
        print('odor done')

        self.waterR = DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(self.waterline))
        self.waterR.low()
        # self.OdorOnCopy = DAQSimpleDOTask('Dev3/port2/line5')
        # self.OdorOnCopy.low()
        self.lickR = DAQSimpleDITask('Dev2_SELECT/port1/line0')
        print('water done')

        # EVENT CODES
        # video recording start / start trial = 101
        # end trial = 100
        # lick on = 11, lick off = 10

        # contingency reward odor on = 131, off = 130, water on = 51, right water off = 50
        # contingency no reward odor on = 141, off = 140, water on = 61, right water off = 60
        # non-contingency reward odor on = 151, off = 150, water on = 71, right water off = 70
        # non-contingency no reward odor on = 161, off = 160, water on = 81, right water off = 80

        # create excel workbook

        self.wb = Workbook()
        self.ws = self.wb.active


        #generate trial type

        # generate trial type
        trialtypes = self.generate_trial()

        self.trialtype = trialtypes
        print('================== Trial Types =================')
        print(trialtypes)

        for t in range(0, self.numtrials):
            print('================================================')

            print('trial number: ', t)
            print()

            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=101)
            self.check_licking_1spout(self.duration_rec_on_before)

#           main training program
            self.run_trial_type(int(trialtypes[t]),t)


            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=100)
            self.wb.save(self.filename)

        odors_cue.initiate()
        odors_cue.close()
        self.waterR.low()
        self.waterR.close()
        logging.basicConfig(filename=logname, level=logging.DEBUG)
        logging.info(self.__dict__)

        print('FINISHED ASSOCIATION TRAINING')




    def check_licking_1spout(self, interval,check_action=False):

        checkperiod = 0.01
        timeout = time.time() + interval
        reward_on = True
        right_lick_last = 0

        count = 0
        while time.time() < timeout:
            right_lick = self.lickR.read()

            if right_lick != right_lick_last:
                if right_lick:

                    print('Lick')
                    d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
                    d = self.ws.cell(row=self.ws.max_row, column=2, value=11)
                # self.save_training()
                    if check_action:
                        count += 1
                else:

                    d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                    d = self.ws.cell(row=self.ws.max_row, column=2, value=10)
            else:
                pass
            right_lick_last = right_lick
            time.sleep(checkperiod)
        if check_action and count >= self.licknum:

            print('licking activate reward')
        elif check_action and count < self.licknum:
            print('not enough licking')
            reward_on = False
        return reward_on

    def run_trial_type(self,types,t):
        odor_on = False
        reward_on = False
        if types == 0:
            print('non-contingency odor reward trial ' + str(int(self.counter[types])))

            # contingency odor comes

            r_code = [151, 150]
            w_code = [51, 50]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_action+self.duration_action_window)
            reward_on = True
            self.run_reward_module(reward_on, w_code)


        elif types == 1:

            print('contingency odor reward trial ' + str(int(self.counter[types])))


            odor_on = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_action)
            reward_on = self.check_licking_1spout(self.duration_action_window, self.operant)
            self.run_reward_module(reward_on, w_code)
            self.check_licking_1spout(self.duration_rec_on_after)
            self.check_licking_1spout(self.duration_ITI[t])
        elif types == 2:
            print(' non-contingency no reward trial ' + str(int(self.counter[types])))
            # odor false: control odor comes
            w_code = [51, 50]

            self.check_licking_1spout(self.duration_odor_on)
            self.check_licking_1spout(self.duration_odor_to_action+self.duration_action_window)
            self.run_reward_module(reward_on, w_code)
        elif types == 3:
            print('contingency odor no reward trial ' + str(int(self.counter[types])))
            odor_on = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_action)  ### can be a problem
            self.check_licking_1spout(self.duration_action_window)
            self.run_reward_module(reward_on, w_code)
            self.check_licking_1spout(self.duration_rec_on_after)
            self.check_licking_1spout(self.duration_ITI[t])

        self.counter[types] += 1

        self.wb.save(self.filename)

    def run_odor_module(self,odor_on, r_code):
        if odor_on:
            print('opening odor port')
            self.reward_odor.high()

            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[0])
            # self.save_training()

            self.check_licking_1spout(self.duration_odor_on)

            print('closing odor port')
            self.reward_odor.low()

            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[1])

        else:

            self.check_licking_1spout(self.duration_odor_on)




    def run_reward_module(self ,reward_on, w_code):
        if reward_on:

            # modify! give water if licks three times within 1 s

            print('opening water valve')
            self.waterR.high()
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[0])

            self.check_licking_1spout(self.duration_water_large)  # this parameter hasn't een defined

            print('closing water valve')
            self.waterR.low()
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[1])


        else:

            self.check_licking_1spout(self.duration_water_large)


#     def camera_action(self):
#         '''
#         format the image properly
#         '''
#         try:
#             wide_image = self.wide_cam.read()
#             wide_image_data = self.wide_cam.to_numpy(wide_image)
#             self.wide_disp_queue.put(wide_image_data)
#
#             if self.settings.save_video.value() and self.settings.in_trial.value() :
#                 self.recorder.save_frame(self.settings.filename.value(),wide_image)
#
#         except Exception as ex:
#             print('Error : %s' % ex)

    def save_training(self):
        with open(self.filename, 'wb') as output:
            pickle.dump(self, output, pickle.HIGHEST_PROTOCOL)




test = SelinaTraining()
print('start')
test.run()


