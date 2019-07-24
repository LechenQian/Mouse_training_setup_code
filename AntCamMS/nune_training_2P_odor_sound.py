'''
Created on Mar 26, 2018

@author: Hao Wu
'''


from ScopeFoundry import Measurement
from ScopeFoundry.measurement import MeasurementQThread
from ScopeFoundry.helper_funcs import sibling_path, load_qt_ui_file
from ScopeFoundry import h5_io
import pyqtgraph as pg
import numpy as np
import random
import winsound
from scipy import ndimage
import time
from numpy.random import rand
import PySpin
from qtpy import QtCore
from qtpy.QtCore import QObject
import os
import queue
from AntCamHW.daq_do.daq_do_dev import DAQSimpleDOTask
from AntCamHW.daq_di.daq_di_dev import DAQSimpleDITask
from time import sleep
from openpyxl import Workbook


class SubMeasurementQThread(MeasurementQThread):
    '''
    Sub-Thread for different loops in measurement
    '''

    def __init__(self, run_func, parent=None):
        '''
        run_func: user-defined function to run in the loop
        parent = parent thread, usually None
        '''
        super(MeasurementQThread, self).__init__(parent)
        self.run_func = run_func
        self.interrupted = False
  
    def run(self):
        while not self.interrupted:
            self.run_func()
            if self.interrupted:
                break
            
    @QtCore.Slot()
    def interrupt(self):
        self.interrupted = True


class NuneTraining(Measurement):
    
    # this is the name of the measurement that ScopeFoundry uses 
    # when displaying your measurement and saving data related to it    
    name = "nune_training"
    interrupt_subthread = QtCore.Signal(())
    
    def setup(self):
        """
        Runs once during App initialization.
        This is the place to load a user interface file,
        define settings, and set up data structures. 
        """
        
        # Define ui file to be used as a graphical interface
        # This file can be edited graphically with Qt Creator
        # sibling_path function allows python to find a file in the same folder
        # as this python module
        self.ui_filename = sibling_path(__file__, "ant_watch_plot.ui")
        
        #Load ui file and convert it to a live QWidget of the user interface
        self.ui = load_qt_ui_file(self.ui_filename)

        # Measurement Specific Settings
        # This setting allows the option to save data to an h5 data file during a run
        # All settings are automatically added to the Microscope user interface
        self.settings.New('save_h5', dtype=bool, initial=True)
        self.settings.New('save_video', dtype = bool, initial = False)
        
        # x and y is for transmitting signal
        # self.settings.New('x',dtype = float, initial = 32, ro = True, vmin = 0, vmax = 63.5)
        # self.settings.New('y',dtype = float, initial = 32, ro = True, vmin = 0, vmax = 63.5)
        
        # added by Nune
        self.settings.New('filename', dtype=str, initial='trial')
        self.settings.New('in_trial', dtype = bool, initial = False)
        self.settings.New('view_only', dtype = bool, initial = False)
        self.settings.New('lick_status', dtype = int, initial = 0)
        self.settings.New('play_frequency',dtype=int,initial=0)
        
        # Define how often to update display during a run
        self.display_update_period = 0.04
        
        
        # Convenient reference to the hardware used in the measurement
        self.wide_cam = self.app.hardware['wide_cam']
        self.recorder = self.app.hardware['flirrec']
        
        #setup experiment condition
        self.wide_cam.settings.frame_rate.update_value(8)
        self.wide_cam.read_from_hardware()

    def setup_figure(self):
        """
        Runs once during App initialization, after setup()
        This is the place to make all graphical interface initializations,
        build plots, etc.
        """
        # connect ui widgets to measurement/hardware settings or functions
        self.ui.start_pushButton.clicked.connect(self.start)
        self.ui.interrupt_pushButton.clicked.connect(self.interrupt)
        
        # Set up pyqtgraph graph_layout in the UI
        self.wide_cam_layout=pg.GraphicsLayoutWidget()
        self.ui.wide_cam_groupBox.layout().addWidget(self.wide_cam_layout)
        
        #create camera image graphs
        self.wide_cam_view=pg.ViewBox()
        self.wide_cam_layout.addItem(self.wide_cam_view)
        self.wide_cam_image=pg.ImageItem()
        self.wide_cam_view.addItem(self.wide_cam_image)

        #counter used for reducing refresh rate
        self.wide_disp_counter = 0
        
    def update_display(self):
        """
        Displays (plots) the numpy array self.buffer. 
        This function runs repeatedly and automatically during the measurement run.
        its update frequency is defined by self.display_update_period
        """
        
        # check availability of display queue of the wide camera
        if not hasattr(self,'wide_disp_queue'):
            pass
        elif self.wide_disp_queue.empty():
            pass
        else:
            try:
                wide_disp_image = self.wide_disp_queue.get()
                if type(wide_disp_image) == np.ndarray:
                    if wide_disp_image.shape == (self.wide_cam.settings.height.value(),self.wide_cam.settings.width.value()):
                        try:
                            self.wide_cam_image.setImage(wide_disp_image)
                        except Exception as ex:
                            print('Error: %s' % ex)
            except Exception as ex:
                print("Error: %s" % ex)
        

    def run(self):
        
        """
        Runs when measurement is started. Runs in a separate thread from GUI.
        It should not update the graphical interface directly, and should only
        focus on data acquisition.
        """
#         # first, create a data file
#         if self.settings['save_h5']:
#             # if enabled will create an HDF5 file with the plotted data
#             # first we create an H5 file (by default autosaved to app.settings['save_dir']
#             # This stores all the hardware and app meta-data in the H5 file
#             self.h5file = h5_io.h5_base_file(app=self.app, measurement=self)
#             
#             # create a measurement H5 group (folder) within self.h5file
#             # This stores all the measurement meta-data in this group
#             self.h5_group = h5_io.h5_create_measurement_group(measurement=self, h5group=self.h5file)
#             
#             # create an h5 dataset to store the data
#             self.buffer_h5 = self.h5_group.create_dataset(name  = 'buffer', 
#                                                           shape = self.buffer.shape,
#                                                           dtype = self.buffer.dtype)
        
        # We use a try/finally block, so that if anything goes wrong during a measurement,
        # the finally block can clean things up, e.g. close the data file object.
        
        #self.wide_cam._dev.set_buffer_count(500)
    
        if self.settings.save_video.value():
            save_dir = self.app.settings.save_dir.value()
            data_path = os.path.join(save_dir,self.app.settings.sample.value())
            try:
                os.makedirs(data_path)
            except OSError:
                print('directory already exists, writing to existing directory')
             
            self.recorder.settings.path.update_value(data_path)
         
            #frame_rate = self.wide_cam.settings.frame_rate.value()
            #self.recorder.create_file('wide_mov',frame_rate)
#         
#         # create a subthread and connect it to the interrupt subthread signal
#         self.wide_disp_queue = queue.Queue(1000)
#         self.camera_thread = SubMeasurementQThread(self.camera_action)
#         self.interrupt_subthread.connect(self.camera_thread.interrupt)
#         #start camera
#         self.wide_cam.start()
#          
#         #start camera subthread
#         self.camera_thread.start()
       
        # initiate all the odor solenoids
        odor0 = DAQSimpleDOTask('Dev3/port0/line0')
        clean0 = DAQSimpleDOTask('Dev3/port0/line1')
        odor1 = DAQSimpleDOTask('Dev3/port0/line2')
        clean1 = DAQSimpleDOTask('Dev3/port0/line3')
        odor2 = DAQSimpleDOTask('Dev3/port0/line4')
        clean2 = DAQSimpleDOTask('Dev3/port0/line5')
        odor3 = DAQSimpleDOTask('Dev3/port0/line6')
        clean3 = DAQSimpleDOTask('Dev3/port0/line7')
        odor4 = DAQSimpleDOTask('Dev3/port1/line0')
        clean4 = DAQSimpleDOTask('Dev3/port1/line1')
        odor5 = DAQSimpleDOTask('Dev3/port1/line2')
        clean5 = DAQSimpleDOTask('Dev3/port1/line3')
        odor6 = DAQSimpleDOTask('Dev3/port1/line4')
        clean6 = DAQSimpleDOTask('Dev3/port1/line5')
        odor7 = DAQSimpleDOTask('Dev3/port1/line6')
        clean7 = DAQSimpleDOTask('Dev3/port1/line7')

# #         # D2-8
#         reward_odor = odor5
#         reward_odor_clean = clean5
#         punish_odor = odor4
#         punish_odor_clean = clean4
#         reward_sound = 1
#         punish_sound = 2
#         events_filename = 'd2-8_7-11-2019_odor_sound.xlsx'
#         #events_filename = 'test.xlsx'
        
#         # D2-9
#         reward_odor = odor4
#         reward_odor_clean = clean4
#         punish_odor = odor5
#         punish_odor_clean = clean5
#         reward_sound = 1
#         punish_sound = 2
#         events_filename = 'd2-9_7-11-2019_odor_sound.xlsx'
# 
#         # D2-10
#         reward_odor = odor5
#         reward_odor_clean = clean5
#         punish_odor = odor4
#         punish_odor_clean = clean4
#         reward_sound = 2
#         punish_sound = 1
#         events_filename = 'd2-10_7-11-2019_odor_sound.xlsx'        
               
        # D1-7 rewards: odor4, sound 1
#         reward_odor = odor4
#         reward_odor_clean = clean4
#         punish_odor = odor5
#         punish_odor_clean = clean5
#         reward_sound = 1
#         punish_sound = 2
#         events_filename = 'd1-7_7-12-2019_odor_sound.xlsx'
#         events_filename = 'test.xlsx'
        
#         # D1-12 rewards: odor5, sound 1
#         reward_odor = odor5
#         reward_odor_clean = clean5
#         punish_odor = odor4
#         punish_odor_clean = clean4
#         reward_sound = 1
#         punish_sound = 2
#         events_filename = 'd1-12_7-11-2019_odor_sound.xlsx'
        
        # D1-13 rewards: odor4, sound 2
#         reward_odor = odor4
#         reward_odor_clean = clean4
#         punish_odor = odor5
#         punish_odor_clean = clean5
#         reward_sound = 2
#         punish_sound = 1
#         events_filename = 'd1-13_7-11-2019_odor_sound.xlsx'
        
        # D1-14 rewards: odor5, sound 2
        reward_odor = odor5
        reward_odor_clean = clean5
        punish_odor = odor4
        punish_odor_clean = clean4
        reward_sound = 2
        punish_sound = 1
        events_filename = 'd1-14_7-11-2019_odor_sound.xlsx'

        
        odor0.low()
        odor1.low()
        odor2.low()
        odor3.low()
        odor4.low()
        odor5.low()
        odor6.low()
        odor7.low()
        
        clean0.low()
        clean1.low()
        clean2.low() 
        clean3.high()
        clean4.high()
        clean5.high()
        clean6.high()
        clean7.high()
        
        odor0.close()
        odor1.close()
        odor2.close()
        odor3.close()
        odor6.close()
        odor7.close()
        
        clean0.close()
        clean1.close()
        clean2.close() 
        clean3.close()
        clean6.close() 
        clean7.close()    
        
        
        LED = DAQSimpleDOTask('Dev3/port2/line2')
        LED.low()
        self.OdorOnCopy = DAQSimpleDOTask('Dev3/port2/line5')
        self.OdorOnCopy.low()
        waterR = DAQSimpleDOTask('Dev3/port2/line0')
        waterR.low()
        airpuff1 = DAQSimpleDOTask('Dev3/port2/line3')
        airpuff1.low()
        airpuff2 = DAQSimpleDOTask('Dev3/port2/line6')
        airpuff2.low()
        self.lickR = DAQSimpleDITask('Dev3/port2/line4')
        
        # EVENT CODES
        # video recording start / start trial = 101
        # lick on = 11, lick off = 10
        # right water on = 51, right water off = 50
        # airpuff on = 81, off = 80
        
        # reward odor on = 131, off = 130
        # punish odor on = 141, off = 140
        # reward sound on = 151, off = 150
        # punish sound on = 161, off = 160
        
        #create excel workbook
        self.wb = Workbook()
        self.ws = self.wb.active
        
        numtrials = 120
        a = np.empty(int(numtrials/4))
        a.fill(2)
        b = np.empty(int(numtrials/4))
        b.fill(3)
        trialtypes = np.concatenate((np.ones((int(numtrials/4))),np.zeros((int(numtrials/4))),a,b))
        random.shuffle(trialtypes)
        print(trialtypes)
        
#         # arrays to randomize skipped trials for each trial type
#         num_to_skip = int((numtrials/5)*0.1)
#         skip0 = np.concatenate((np.zeros(int(numtrials/10)-num_to_skip),np.ones(num_to_skip)))
#         random.shuffle(skip0)
#         skip0 = np.concatenate((np.zeros(int(numtrials/10)),skip0))
#         skip1 = np.concatenate((np.zeros(int(numtrials/10)-num_to_skip),np.ones(num_to_skip)))
#         random.shuffle(skip1)
#         skip1 = np.concatenate((np.zeros(int(numtrials/10)),skip1))
#         skip2 = np.concatenate((np.zeros(int(numtrials/10)-num_to_skip),np.ones(num_to_skip)))
#         random.shuffle(skip2)
#         skip2 = np.concatenate((np.zeros(int(numtrials/10)),skip2))
#         skip3 = np.concatenate((np.zeros(int(numtrials/10)-num_to_skip),np.ones(num_to_skip)))
#         random.shuffle(skip3)
#         skip3 = np.concatenate((np.zeros(int(numtrials/10)),skip3))
        

        # counters for each trial type
        h = 0
        i = 0
        j = 0
        k = 0
        
        duration_rec_off = 6.5
        duration_rec_on_before = 4
        duration_odor_to_outcome = 1.3
        duration_water_large = 0.2
        duration_airpuff = 0.2
        duration_rec_on_after = 8
        
        sound_thread = SubMeasurementQThread(self.sound_action)
        self.interrupt_subthread.connect(sound_thread.interrupt)
        sound_thread.start()
            
        for t in range(0,numtrials):
            
            if bool(trialtypes[t]):
                self.settings.filename.update_value('reward_trial' + str(t))
            else:
                self.settings.filename.update_value('punish_trial' + str(t))
                    
            self.recorder.create_file(self.settings.filename.value(),self.wide_cam.settings.frame_rate.value())
            
            #print('turn on LED')
            LED.high()
            self.check_licking_1spout(1)
            
            #self.settings.save_video.update_value(True):
            print(t)
            self.settings.in_trial.update_value(True)
            d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=101)
            self.check_licking_1spout(duration_rec_on_before)
                
            if trialtypes[t] == 0:
                print('odor reward trial ' + str(h))
                print('opening odor port')
                reward_odor.high()
                self.OdorOnCopy.high()
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=131)
                
                reward_odor_clean.low() 
                self.check_licking_1spout(duration_odor_to_outcome)
                
                print('opening water valve')
                waterR.high()
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=51)
                self.check_licking_1spout(duration_water_large)
                
                print('closing odor port')
                reward_odor.low()
                self.OdorOnCopy.low()
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=130)
                reward_odor_clean.high()
                            
                waterR.low()
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=50)
                
                h += 1
                
                self.wb.save(events_filename)
                
            elif trialtypes[t] == 1:
                print('punish odor trial ' + str(i))
                print('opening odor port')
                punish_odor.high()
                self.OdorOnCopy.high()
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=141)
                
                punish_odor_clean.low()
                self.check_licking_1spout(duration_odor_to_outcome)
                
                print('delivering airpuff')
                airpuff1.high()
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=81)

                self.check_licking_1spout(duration_airpuff)
            
                print('closing odor port')
                punish_odor.low()
                self.OdorOnCopy.low()
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=140)
                punish_odor_clean.high()
                
                airpuff1.low()
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=80)
                
                i +=1
                
                self.wb.save(events_filename)
                
            elif trialtypes[t] == 2:
                print('sound reward trial ' + str(j))
                print('playing reward sound')
                
                self.settings.play_frequency.update_value(reward_sound)
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=151)
                
                self.check_licking_1spout(duration_odor_to_outcome)
                
                print('opening water valve')
                waterR.high()
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=51)
                self.check_licking_1spout(duration_water_large)
               
                waterR.low()
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=50)
                
                j += 1
                
                self.wb.save(events_filename)
                
            else:
                print('sound punish trial ' + str(k))
                print('playing punish sound')

                self.settings.play_frequency.update_value(punish_sound)
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=161)

                self.check_licking_1spout(duration_odor_to_outcome)
                
                print('delivering airpuff')
                airpuff1.high()
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=81)

                self.check_licking_1spout(duration_airpuff)
                
                airpuff1.low()
                d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                d = self.ws.cell(row=self.ws.max_row, column=2, value=80)
                
                k += 1
                
                self.wb.save(events_filename)
            

            self.check_licking_1spout(duration_rec_on_after)
            self.settings.in_trial.update_value(False)
            #self.settings.save_video.update_value(False):
            LED.low()
            
            self.wb.save(events_filename)
    
            self.check_licking_1spout(duration_rec_off)
    
            if self.interrupt_measurement_called:
                #tell subtherad to stop
                self.interrupt_subthread.emit()
                break
            
        reward_odor.close()
        reward_odor_clean.close()
        punish_odor.close()
        punish_odor_clean.close()
        waterR.close()
        waterR.close()
        airpuff1.close()
        airpuff2.close()
        LED.close()
        self.OdorOnCopy.close()
        print('FINISHED ASSOCIATION TRAINING')
        
        #sound_thread.stop()
        del sound_thread
        
        if self.settings.save_video.value():
            self.recorder.close()           
                 
    
    def check_licking_1spout(self,interval):
        
        checkperiod = 0.02
        timeout = time.time() + interval
        
        right_lick_last = 0
        while time.time() < timeout:
            right_lick = self.lickR.read()
              
            if right_lick != right_lick_last:
                if right_lick:
                    self.settings.lick_status.update_value(11)
                    print('Lick')
                    d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                    d = self.ws.cell(row=self.ws.max_row, column=2, value=11)
                else:
                    self.settings.lick_status.update_value(10)
                    d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                    d = self.ws.cell(row=self.ws.max_row, column=2, value=10)
            else:
                self.settings.lick_status.update_value(0)

            right_lick_last = right_lick
            time.sleep(checkperiod)


    def sound_action(self):
        checkperiod = 0.02
        time.sleep(checkperiod)
        if self.settings.play_frequency.value()==1:
            self.settings.play_frequency.update_value(0)
            self.OdorOnCopy.high()
            winsound.Beep(5000,1500)
            #winsound.PlaySound('highsoundstimulus.wav', winsound.SND_ALIAS)
            self.OdorOnCopy.low()
        elif self.settings.play_frequency.value()==2:
            self.settings.play_frequency.update_value(0)
            self.OdorOnCopy.high()
            #winsound.PlaySound('lowsoundstimulus.wav', winsound.SND_ALIAS)
            winsound.Beep(12000,1500)
            self.OdorOnCopy.low()

            
    
#     def camera_action(self):
#         '''
#         format the image properly
#         '''
#         try:
#             wide_image = self.wide_cam.read()
#             wide_image_data = self.wide_cam.to_numpy(wide_image)
#             self.wide_disp_queue.put(wide_image_data)
#             
#             if self.settings.save_video.value() and self.settings.in_trial.value() :
#                 self.recorder.save_frame(self.settings.filename.value(),wide_image)
# 
#         except Exception as ex:
#             print('Error : %s' % ex)