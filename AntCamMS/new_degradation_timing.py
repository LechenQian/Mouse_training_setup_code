'''
Created on July 17, 2019
full version taht covers all the behavioral paradigm I'm running.
@author: Selina Qian
'''
from ScopeFoundry import Measurement
import datetime
import numpy as np
import random
import pickle
import time
from AntCamHW.daq_do.daq_do_dev import DAQSimpleDOTask
from AntCamHW.daq_di.daq_di_dev import DAQSimpleDITask
from openpyxl import Workbook
import os
import logging

class OdorGen(object):
    def __init__(self,odorindex):
        self.odorindex = odorindex
        self.odors_DAQ = []


    def assign_odor(self):
        # initiate all the odor solenoids
        for item in self.odorindex:
            self.odors_DAQ.append(DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(item)))
        print('Odor {} has been properly assigned'.format(self.odorindex))

    def set_rewardodor(self, index: list):
        reward_odor = self.odors_DAQ[index[0]]
        non_reward_odor = self.odors_DAQ[index[1]]
        if len(self.odorindex) == 3:
            control_odor = self.odors_DAQ[index[2]]

            print('reward odor, unrewarded odor and control odor loaded')
            return reward_odor, non_reward_odor, control_odor
        else:
            print('reward odor, unrewarded odor loaded')
            return reward_odor,  non_reward_odor

    def initiate(self):
        for odor in self.odors_DAQ:
            odor.low()
        print('Odor initiation: status low')

    def close(self):
        for odor in self.odors_DAQ:
            odor.close()
        print('Connection has been closed')


class SelinaTraining(Measurement):
    def __init__(self):
        # please change this according to mouse
        self.mouse = 'C42'  #OT-GC-2

        self.phase = 'rep_deg' #'cond', 'rep_deg'
        self.condition = 'Pav'
        self.numtrials = 100

        self.list = [7, 6, 5]
        self.events_path = "C:/Users/MurthyLab/Desktop/Selina_lab_computer_data/experiment_data_2021_2_{0}/{1}/".format(self.condition,self.mouse)

        self.odor_index = [1,0,2] #odor list index position 0 is reward, 1 is unrewarded, 2 is control c odor; so here, 7 is unrewarded, 6 is reward and 5 is c odor
        if self.condition == 'Operant':
            self.operant = True
            self.licknum = 1
        else:
            self.operant = False
            self.licknum = 0


        # pre training
        self.p_go = 0.5#0.5
        self.p_reward_go = 0.9
        self.p_no_go = 0.5#0.2
        self.p_empty = 0#0.3
        if self.phase == 'cond':
            self.p_reward_empty = 0  # 0.3
        else: # 'deg','C-control','close','far'
            self.p_reward_empty = 0.75


        self.counter = np.zeros(9)

        self.duration_rec_on_before = 10
        self.preceding_interval = 1
        self.bef_unpred_rew = self.duration_rec_on_before-self.preceding_interval
        self.duration_odor_on = 1
        self.duration_odor_to_action = 0
        self.duration_action_window = 2.5
        self.duration_water_large = 0.06 #0.1 if one tube, 0.18 if two tube
        self.duration_rec_on_after = 10 #close 1, cond & deg 3
        self.duration_ITI = np.random.exponential(6, size=self.numtrials)
        self.events_filename = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M") + '{}_{}.xlsx'.format(self.phase,
                                                                                                        self.preceding_interval)

        self.waterline = 0
        self.filename = self.events_path + self.events_filename

    def run(self):
        try:
            os.makedirs(self.events_path)
        except OSError:
            print("The directory %s existed" % self.events_path)
        else:
            print("Successfully created the directory %s " % self.events_path)
        logname = self.filename[0:-5] + '.log'
        logging.basicConfig(filename=logname, level=logging.DEBUG)
        logging.info(self.__dict__)
        odors_cue = OdorGen(self.list)
        odors_cue.assign_odor()
        self.reward_odor, self.non_reward_odor, self.control_odor = odors_cue.set_rewardodor(index=self.odor_index)
        odors_cue.initiate()
        # odors_cue.odors_DAQ[i]
        print('odor done')

        self.waterR = DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(self.waterline))
        self.waterR.low()
        # self.OdorOnCopy = DAQSimpleDOTask('Dev3/port2/line5')
        # self.OdorOnCopy.low()
        self.lickR = DAQSimpleDITask('Dev2_SELECT/port1/line0')
        print('water done')

        # EVENT CODES
        # video recording start / start trial = 101
        # end trial = 100
        # lick on = 11, lick off = 10

        # reward odor on = 131, off = 130, water on = 51, water off = 50
        # no reward odor on = 141, off = 140
        # c control odor on = 161, off = 160,  water on = 51, water off = 50

        # create excel workbook

        self.wb1 = Workbook()
        self.ws1 = self.wb1.active


        #generate trial type

        # generate trial type
        trialtypes = np.zeros(self.numtrials)
        for i in range(int(self.numtrials/20)):

            if self.phase == 'cond':
                train_go = np.zeros(int(round(20 * self.p_go * self.p_reward_go)))  # code 0
                train_nogo = np.ones(int(20 * self.p_no_go))  # code 1
                temp_comb1 = np.concatenate((train_go, train_nogo))
                train_go_omission = np.ones(int(round(20 * self.p_go * (1 - self.p_reward_go)))) * 3  # code 3
                temp_comb1 = np.concatenate((temp_comb1, train_go_omission))


            elif self.phase == 'rep_deg':
                train_go_unpred = np.ones(int(round(20 * self.p_go * self.p_reward_go))) * 2  # code 2
                train_nogo = np.ones(int(20 * self.p_no_go))  # code 1
                temp_comb1 = np.concatenate((train_go_unpred, train_nogo))
                train_go_omission = np.ones(int(round(20 * self.p_go * (1 - self.p_reward_go)))) * 3  # code 3
                temp_comb1 = np.concatenate((temp_comb1, train_go_omission))


            temp_comb2 = temp_comb1.copy()
            random.shuffle(temp_comb2)
            trialtypes[20*i:20*(i+1)] = temp_comb2
        self.trialstype = trialtypes
        print('================== Trial Types =================')
        print(trialtypes)
        for t in range(0, self.numtrials):
            print('================================================')
            print('trial number: ', t)

            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=101) #trial start
            d = self.ws1.cell(row=self.ws1.max_row, column=3, value= 'trial{}'.format(int(trialtypes[t]))) #trial type
            # code: 0--go w rward; 1--no go; 2--empty; 3--go w/o reward; 4 --unpred water; 5 -- c control odor； 6 -- c control odor w/o


            #           main training program
            self.run_trial_type(int(trialtypes[t]))

            self.check_licking_1spout(self.duration_rec_on_after)
            # self.settings.save_video.update_value(False):
            self.check_licking_1spout(self.duration_ITI[t])
            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=100) #end
            self.wb1.save(self.filename)
        odors_cue.initiate()
        odors_cue.close()
        self.waterR.low()
        self.waterR.close()
        print('FINISHED ASSOCIATION TRAINING')

    def check_licking_1spout(self, interval,check_action=False):
        checkperiod = 0.005
        timeout = time.time() + interval
        reward_on = True # this is used for Pav or Operant
        right_lick_last = 0
        count = 0
        while time.time() < timeout:
            right_lick = self.lickR.read()
            if right_lick != right_lick_last:
                if right_lick:
                    print('Lick')
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)
                # self.save_training()
                    if check_action:
                        count += 1
                else:
                    d = self.ws1.cell(row=(self.ws1.max_row+1), column=1, value=time.clock())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
            else:
                pass
            right_lick_last = right_lick
            time.sleep(checkperiod)
        if check_action and count >= self.licknum:

            print('licking activate reward')
        elif check_action and count < self.licknum:
            print('not enough licking')
            reward_on = False
        return reward_on

    def run_trial_type(self, types):
        odor_on = False
        reward_on = False
        is_control = False
        if types == 0:
            print('go trial ' + str(int(self.counter[types])))
            odor_on = True # contingency odor comes
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action)
            reward_on = self.check_licking_1spout(self.duration_action_window, self.operant)  ### can be a problem
            self.run_reward_module(reward_on, w_code)

        elif types == 1:
            print('no go trial ' + str(int(self.counter[types])))

            odor_on = True
            is_go = False
            r_code = [141, 140]
            w_code = [61, 60]
            self.check_licking_1spout(self.duration_rec_on_before)
            self.run_odor_module(odor_on,is_go, is_control, r_code)


        elif types == 2:
            print('degradation  trial ' + str(int(self.counter[types])))
            # odor false: control odor comes
            w_code = [71, 70]# not used
            odor_on = True  # contingency odor comes
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            reward_on = True
            self.check_licking_1spout(self.bef_unpred_rew)
            self.run_reward_module(reward_on, w_code)
            self.check_licking_1spout(self.preceding_interval)
            self.run_odor_module(odor_on, is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action)
            reward_on = self.check_licking_1spout(self.duration_action_window, self.operant)  ### can be a problem
            self.run_reward_module(reward_on, w_code)

        elif types == 3:
            print('probe trial ' + str(int(self.counter[types])))
            odor_on = True # contingency odor comes
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.check_licking_1spout(self.duration_rec_on_before)
            self.run_odor_module(odor_on, is_go, is_control, r_code)


        self.counter[types] += 1

        self.wb1.save(self.filename)

    def run_odor_module(self, odor_on, is_go, is_control, r_code):
        if odor_on:
            print('opening odor port')
            if is_go and not is_control:
                self.reward_odor.high()
            elif not is_go and not is_control:
                self.non_reward_odor.high()
            elif is_control:
                self.control_odor.high()
            # self.OdorOnCopy.high()  # ？？？
            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=r_code[0])
            # self.save_training()
            self.check_licking_1spout(self.duration_odor_on)
            print('closing odor port')
            if is_go and not is_control:
                self.reward_odor.low()
            elif not is_go and not is_control:
                self.non_reward_odor.low()
            elif is_control:
                self.control_odor.low()

            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=r_code[1])
        else:
            self.check_licking_1spout(self.duration_odor_on)

    def run_reward_module(self, reward_on, w_code):
        if reward_on:
            # modify! give water if licks three times within 1 s
            print('opening water valve')
            self.waterR.high()
            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=w_code[0])
            self.check_licking_1spout(self.duration_water_large)  # this parameter hasn't een defined
            print('closing water valve')
            self.waterR.low()
            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=w_code[1])
        else:
            self.check_licking_1spout(self.duration_water_large)


#     def camera_action(self):
#         '''
#         format the image properly
#         '''
#         try:
#             wide_image = self.wide_cam.read()
#             wide_image_data = self.wide_cam.to_numpy(wide_image)
#             self.wide_disp_queue.put(wide_image_data)
#
#             if self.settings.save_video.value() and self.settings.in_trial.value() :
#                 self.recorder.save_frame(self.settings.filename.value(),wide_image)
#
#         except Exception as ex:
#             print('Error : %s' % ex)

    def save_training(self):
        with open(self.filename, 'wb') as output:
            pickle.dump(self, output, pickle.HIGHEST_PROTOCOL)




test = SelinaTraining()
print('start')
test.run()


