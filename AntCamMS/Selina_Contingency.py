'''
Created on July 17, 2019

@author: Hao Wu, modified by Selina Qian
'''
from ScopeFoundry import Measurement
from ScopeFoundry.measurement import MeasurementQThread
from ScopeFoundry.helper_funcs import sibling_path, load_qt_ui_file
from ScopeFoundry import h5_io
import pyqtgraph as pg
import numpy as np
import random

from scipy import ndimage
import time
from numpy.random import rand
# import PySpin
from qtpy import QtCore
from qtpy.QtCore import QObject
import os
import queue
from AntCamHW.daq_do.daq_do_dev import DAQSimpleDOTask
from AntCamHW.daq_di.daq_di_dev import DAQSimpleDITask
import cv2
from openpyxl import Workbook


class SubMeasurementQThread(MeasurementQThread):
    '''
    Sub-Thread for different loops in measurement
    '''

    def __init__(self, run_func, parent=None):
        '''
        run_func: user-defined function to run in the loop
        parent = parent thread, usually None
        '''
        super(MeasurementQThread, self).__init__(parent)
        self.run_func = run_func
        self.interrupted = False

    def run(self):
        while not self.interrupted:
            self.run_func()
            if self.interrupted:
                break

    @QtCore.Slot()
    def interrupt(self):
        self.interrupted = True


class SelinaTraining(Measurement):
    # this is the name of the measurement that ScopeFoundry uses
    # when displaying your measurement and saving data related to it
    name = "selina_training"
    interrupt_subthread = QtCore.Signal(())

    # def __init__(self):
    #     super().__init__()
    #     # EVENT CODES
    #     # video recording start / start trial = 101
    #     # lick on = 11, lick off = 10
    #     # right
    #     # airpuff on = 81, off = 80
    #
    #     # contingency reward odor on = 131, off = 130, water on = 51, right water off = 50
    #     # contingency no reward odor on = 141, off = 140, water on = 61, right water off = 60
    #     # non-contingency reward odor on = 151, off = 150, water on = 71, right water off = 70
    #     # non-contingency no reward odor on = 161, off = 160, water on = 81, right water off = 80
    #
    #     self.events_filename = '2019-7-17-test.xlsx'
    #     self.reward_odor_index = 0
    #
    #     self.duration_rec_off = 6.5
    #     self.duration_rec_on_before = 4  # change this to exponential decay
    #     self.duration_odor_to_outcome = 1.3
    #     self.duration_water_large = 0.2
    #     self.duration_rec_on_after = 8
    #     self.duration_odor_on = 0.5
    #
    #     self.lines = [0, 1, 2, 3]
    #     self.counter = np.zeros(len(self.lines))
    #
    #     self.numtrials = 200
    #     self.p_cont_noncont = 0.5
    #     self.p_USwCS = 0.5
    #     self.p_USwoCS = 0.5

    def setup(self):
        """
        Runs once during App initialization.
        This is the place to load a user interface file,
        define settings, and set up data structures.
        """

        # Define ui file to be used as a graphical interface
        # This file can be edited graphically with Qt Creator
        # sibling_path function allows python to find a file in the same folder
        # as this python module
        self.ui_filename = sibling_path(__file__, "ant_watch_plot.ui")

        # Load ui file and convert it to a live QWidget of the user interface
        self.ui = load_qt_ui_file(self.ui_filename)

        # Measurement Specific Settings
        # This setting allows the option to save data to an h5 data file during a run
        # All settings are automatically added to the Microscope user interface
        self.settings.New('save_h5', dtype=bool, initial=True)
        self.settings.New('save_video', dtype=bool, initial=False)

        # x and y is for transmitting signal
        self.settings.New('x', dtype=float, initial=32, ro=True, vmin=0, vmax=63.5)
        self.settings.New('y', dtype=float, initial=32, ro=True, vmin=0, vmax=63.5)

        # added by Nune
        self.settings.New('filename', dtype=str, initial='trial')
        self.settings.New('in_trial', dtype=bool, initial=False)
        self.settings.New('view_only', dtype=bool, initial=False)
        self.settings.New('lick_status', dtype=int, initial=0)

        # added for Selina

        # connect lick indicator to settings
        self.settings.New('lick_on', dtype=bool, initial = False, ro = True)

        # change the lick indicator into blue light
        #self.ui.right_lick_ind_checkBox.setStyleSheet(
        #    'QCheckBox{color:blue;}QCheckBox::indicator:checked{image: url(./icons/c_b.png);}QCheckBox::indicator:unchecked{image: url(./icons/uc_b.png);}')

        # Define how often to update display during a run
        self.display_update_period = 0.01

        # Convenient reference to the hardware used in the measurement
        self.wide_cam = self.app.hardware['HD Webcam C270']
        # self.recorder = self.app.hardware['flirrec']

        # setup experiment condition
        #self.wide_cam.settings.frame_rate.update_value(8)
        self.wide_cam.read_from_hardware()

        self.vc = cv2.VideoCapture(0)
        self.vc.set(cv2.CAP_FFMPEG, True)
        self.vc.set(cv2.CAP_PROP_FPS, 30)

    def setup_figure(self):
        """
        Runs once during App initialization, after setup()
        This is the place to make all graphical interface initializations,
        build plots, etc.
        """
        # connect ui widgets to measurement/hardware settings or functions
        self.ui.start_pushButton.clicked.connect(self.start)
        self.ui.interrupt_pushButton.clicked.connect(self.interrupt)

        # Set up pyqtgraph graph_layout in the UI
        self.wide_cam_layout = pg.GraphicsLayoutWidget()
        self.ui.wide_cam_groupBox.layout().addWidget(self.wide_cam_layout)

        # create camera image graphs
        self.wide_cam_view = pg.ViewBox()
        # rval, frame = self.vc.read()
        self.wide_cam_layout.addItem(self.wide_cam_view)
        self.wide_cam_image = pg.ImageItem()
        self.wide_cam_view.addItem(self.wide_cam_image)

        # counter used for reducing refresh rate
        self.wide_disp_counter = 0

        # connect setting to user interface
        #self.settings.lick_on.connect_to_widget(self.ui.lick_checkBox)

    def update_display(self):
        """
        Displays (plots) the numpy array self.buffer.
        This function runs repeatedly and automatically during the measurement run.
        its update frequency is defined by self.display_update_period
        """

        # check availability of display queue of the wide camera
        try:
            ret, frame = self.vc.read()
            self.wide_cam_image.setImage(np.flipud(np.fliplr(frame[:, :, 2::-1].transpose(1, 0, 2))))
        except Exception as ex:
            print("Error: %s" % ex)

    def run(self):

        """
        Runs when measurement is started. Runs in a separate thread from GUI.
        It should not update the graphical interface directly, and should only
        focus on data acquisition.
        """
        #         # first, create a data file
        #         if self.settings['save_h5']:
        #             # if enabled will create an HDF5 file with the plotted data
        #             # first we create an H5 file (by default autosaved to app.settings['save_dir']
        #             # This stores all the hardware and app meta-data in the H5 file
        #             self.h5file = h5_io.h5_base_file(app=self.app, measurement=self)
        #
        #             # create a measurement H5 group (folder) within self.h5file
        #             # This stores all the measurement meta-data in this group
        #             self.h5_group = h5_io.h5_create_measurement_group(measurement=self, h5group=self.h5file)
        #
        #             # create an h5 dataset to store the data
        #             self.buffer_h5 = self.h5_group.create_dataset(name  = 'buffer',
        #                                                           shape = self.buffer.shape,
        #                                                           dtype = self.buffer.dtype)

        # We use a try/finally block, so that if anything goes wrong during a measurement,
        # the finally block can clean things up, e.g. close the data file object.

        # self.wide_cam._dev.set_buffer_count(500)

        # if self.settings.save_video.value():
        #     save_dir = self.app.settings.save_dir.value()
        #     data_path = os.path.join(save_dir, self.app.settings.sample.value())
        #     try:
        #         os.makedirs(data_path)
        #     except OSError:
        #         print('directory already exists, writing to existing directory')
        #
        #     self.recorder.settings.path.update_value(data_path)

            # frame_rate = self.wide_cam.settings.frame_rate.value()
            # self.recorder.create_file('wide_mov',frame_rate)
        #
        #         # create a subthread and connect it to the interrupt subthread signal
        #         self.wide_disp_queue = queue.Queue(1000)
        #         self.camera_thread = SubMeasurementQThread(self.camera_action)
        #         self.interrupt_subthread.connect(self.camera_thread.interrupt)
        #         #start camera
        #         self.wide_cam.start()
        #
        #         #start camera subthread
        #         self.camera_thread.start()
        self.events_filename = '2019-7-17-test.xlsx'
        self.reward_odor_index = 0

        self.duration_rec_off = 6.5
        self.duration_rec_on_before = 4  # change this to exponential decay
        self.duration_odor_to_outcome = 1.3
        self.duration_water_large = 0.2
        self.duration_rec_on_after = 8
        self.duration_odor_on = 0.5

        self.lines = [0, 1, 2, 3]
        self.counter = np.zeros(4)

        self.numtrials = 200
        self.p_cont_noncont = 0.5
        self.p_USwCS = 0.5
        self.p_USwoCS = 0.5



        odors_cue = OdorGen(self.lines)
        odors_cue.assign_odor()
        self.reward_odor = odors_cue.set_rewardodor(index=self.reward_odor_index)
        odors_cue.initiate()
        # odors_cue.odors_DAQ[i]
        print('odor ready')

        self.waterR = DAQSimpleDOTask('Dev1/port0/line2')
        self.waterR.low()
        # self.OdorOnCopy = DAQSimpleDOTask('Dev3/port2/line5')
        # self.OdorOnCopy.low()
        self.lickR = DAQSimpleDITask('Dev2_SELECT/port1/line0')

        print('water ready')

        # create excel workbook
        self.wb = Workbook()
        self.ws = self.wb.active
        print('book ready')


        #generate trial type

        cont_reward = np.zeros(int(self.numtrials * self.p_cont_noncont * self.p_USwCS))  # code 0
        cont_noreward = np.ones(int(self.numtrials * self.p_cont_noncont * (1 - self.p_USwCS)))  # code 1
        temp_comb1 = np.concatenate((cont_reward, cont_noreward))

        noncont_reward = np.ones(int(self.numtrials * (1 - self.p_cont_noncont) * self.p_USwoCS)) * 2  # code 2
        noncont_noreward = np.ones(int(self.numtrials * (1 - self.p_cont_noncont) * (1 - self.p_USwoCS))) * 3  # code 3
        temp_comb2 = np.concatenate((noncont_noreward, noncont_reward))

        trialtypes = np.concatenate((temp_comb1, temp_comb2))
        random.shuffle(trialtypes)
        print('================== Trial Types =================')
        print(trialtypes)




        for t in range(0, self.numtrials):
            print('================================================')
            print('trial number: ', t)
            print()
            self.settings.in_trial.update_value(True)
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=101)
            # check lisking
            self.check_licking_1spout(self.duration_rec_on_before)

#           main training program
            self.run_trial_type(int(trialtypes[t]))
            # check licking
            self.check_licking_1spout(self.duration_rec_on_after)
            self.settings.in_trial.update_value(False)
            # self.settings.save_video.update_value(False):

            self.wb.save(self.events_filename)

            self.check_licking_1spout(self.duration_rec_off)

            if self.interrupt_measurement_called:
                # tell subtherad to stop
                self.interrupt_subthread.emit()
                break

        odors_cue.initiate()
        odors_cue.close()
        self.waterR.low()
        self.waterR.close()
        print('FINISHED ASSOCIATION TRAINING')


        if self.settings.save_video.value:
            self.recorder.close()


    def check_licking_1spout(self, interval):

        checkperiod = 0.01
        timeout = time.time() + interval
        right_lick_last = 0
        while time.time() < timeout:
            right_lick = self.lickR.read()

            if right_lick != right_lick_last:
                if right_lick:
                    self.settings.lick_status.update_value(11)
                    print('Lick')
                    d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
                    d = self.ws.cell(row=self.ws.max_row, column=2, value=11)
                    self.settings.lick_on.update_value(True)
                else:
                    self.settings.lick_status.update_value(10)
                    d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
                    d = self.ws.cell(row=self.ws.max_row, column=2, value=10)
                    self.settings.lick_on.update_value(False)
            else:
                self.settings.lick_status.update_value(0)

            right_lick_last = right_lick
            time.sleep(checkperiod)

    def run_trial_type(self, types):
        odor_on = False
        reward_on = False
        if types == 0:
            print('contingency reward trial ' + str(int(self.counter[types])))
            print('opening odor port')
            odor_on = True
            reward_on = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_outcome)  ### can be a problem
            self.run_reward_module(reward_on, w_code)

        elif types == 1:
            print('contingency no reward trial ' + str(int(self.counter[types])))
            print('opening odor port')
            odor_on = True
            r_code = [141, 140]
            w_code = [61, 60]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_outcome)  ### can be a problem
            self.run_reward_module(reward_on, w_code)
        elif types == 2:
            print('non-contingency reward trial ' + str(int(self.counter[types])))
            reward_on = True
            r_code = [151, 150]
            w_code = [71, 70]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_outcome)  ### can be a problem
            self.run_reward_module(reward_on, w_code)
        else:
            print('non-contingency no reward trial ' + str(int(self.counter[types])))
            r_code = [161, 160]
            w_code = [71, 70]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_outcome)  ### can be a problem
            self.run_reward_module(reward_on, w_code)

        self.counter[types] += 1

        self.wb.save(self.events_filename)

    def run_odor_module(self, odor_on, r_code):
        if odor_on:
            self.reward_odor.high()
            # self.OdorOnCopy.high()  # ？？？
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[0])

            time.sleep(self.duration_odor_on)

            print('closing odor port')
            self.reward_odor.low()

            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[1])
        else:
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[0])
            time.sleep(self.duration_odor_on)
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[1])

    def run_reward_module(self, reward_on, w_code):
        if reward_on:

            # modify! give water if licks three times within 1 s

            print('opening water valve')
            self.waterR.high()
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[0])
            self.check_licking_1spout(self.duration_water_large)  # this parameter hasn't een defined

            print('closing water valve')
            self.waterR.low()
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[1])

        else:
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[0])
            time.sleep(self.duration_water_large)
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[1])

#     def camera_action(self):
#         '''
#         format the image properly
#         '''
#         try:
#             wide_image = self.wide_cam.read()
#             wide_image_data = self.wide_cam.to_numpy(wide_image)
#             self.wide_disp_queue.put(wide_image_data)
#
#             if self.settings.save_video.value() and self.settings.in_trial.value() :
#                 self.recorder.save_frame(self.settings.filename.value(),wide_image)
#
#         except Exception as ex:
#             print('Error : %s' % ex)


class OdorGen(object):
    def __init__(self,odorindex):
        self.odorindex = odorindex
        self.odors_DAQ = []

    def assign_odor(self):
        # initiate all the odor solenoids

        for item in self.odorindex:
            self.odors_DAQ.append(DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(item)))
        print('Odor {} has been properly assigned'.format(self.odorindex))


    def set_rewardodor(self,index):
        reward_odor = self.odors_DAQ[index]
        print('reward odor is odor {}'.format(index))
        return reward_odor

    def initiate(self):
        for odor in self.odors_DAQ:
            odor.low()
        print('Odor initiation: status low')

    def close(self):
        for odor in self.odors_DAQ:
            odor.close()
        print('Connection has been closed')

