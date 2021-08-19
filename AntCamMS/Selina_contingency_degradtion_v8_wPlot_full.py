'''
Created on July 17, 2019
full version taht covers all the behavioral paradigm I'm running.
@author: Selina Qian
'''
from ScopeFoundry import Measurement
import datetime
import numpy as np
import random
import pickle
import time
from AntCamHW.daq_do.daq_do_dev import DAQSimpleDOTask
from AntCamHW.daq_di.daq_di_dev import DAQSimpleDITask
from openpyxl import Workbook
import os
import logging
import pandas as pd
import matplotlib.pyplot as plt
from collections import OrderedDict
import os
import matplotlib as mpl
import re
import csv


class OdorGen(object):
    def __init__(self,odorindex):
        self.odorindex = odorindex
        self.odors_DAQ = []


    def assign_odor(self):
        # initiate all the odor solenoids
        for item in self.odorindex:
            self.odors_DAQ.append(DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(item)))
        print('Odor {} has been properly assigned'.format(self.odorindex))

    def set_rewardodor(self, index: list):
        reward_odor = self.odors_DAQ[index[0]]
        non_reward_odor = self.odors_DAQ[index[1]]
        if len(self.odorindex) == 3:
            control_odor = self.odors_DAQ[index[2]]

            print('reward odor, unrewarded odor and control odor loaded')
            return reward_odor, non_reward_odor, control_odor
        else:
            print('reward odor, unrewarded odor loaded')
            return reward_odor,  non_reward_odor

    def initiate(self):
        for odor in self.odors_DAQ:
            odor.low()
        print('Odor initiation: status low')

    def close(self):
        for odor in self.odors_DAQ:
            odor.close()
        print('Connection has been closed')


class SelinaTraining(Measurement):
    def __init__(self):
        # please change this according to mouse
        self.mouse = 'C51'  #OT-GC-2

        self.phase = 'cond' #'cond', 'deg','C_control','close','far'
        self.condition = 'Pav'
        self.numtrials = 160

        self.list = [7, 6, 5]
        self.events_path = "C:/Users/MurthyLab/Desktop/Selina_lab_computer_data/experiment_data_2021_2_{0}/{1}/".format(self.condition,self.mouse)
        self.events_filename = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M")+'{}.xlsx'.format(self.phase)
        self.odor_index = [1,0,2] #odor list index position 0 is reward, 1 is unrewarded, 2 is control c odor; so here, 7 is unrewarded, 6 is reward and 5 is c odor
        if self.condition == 'Operant':
            self.operant = True
            self.licknum = 1
        else:
            self.operant = False
            self.licknum = 0


        # pre training
        self.p_go = 0.4#0.5
        self.p_reward_go = 0.75
        self.p_no_go = 0.2#0.2
        self.p_empty = 0.4#0.3
        if self.phase == 'cond':
            self.p_reward_empty = 0  # 0.3
        else: # 'deg','C-control','close','far'
            self.p_reward_empty = 0.75


        self.counter = np.zeros(9)

        self.duration_rec_on_before = 2
        self.duration_odor_on = 1
        self.duration_odor_to_action = 0
        self.duration_action_window = 2.5
        self.duration_water_large = 0.1 #0.1 if one tube, 0.18 if two tube
        self.duration_rec_on_after = 8 #close 1, cond & deg 3
        self.duration_ITI = np.random.exponential(2, size=self.numtrials)

        self.waterline = 0
        self.filename = self.events_path + self.events_filename

    def run(self):
        try:
            os.makedirs(self.events_path)
        except OSError:
            print("The directory %s existed" % self.events_path)
        else:
            print("Successfully created the directory %s " % self.events_path)
        logname = self.filename[0:-5] + '.log'
        logging.basicConfig(filename=logname, level=logging.DEBUG)
        logging.info(self.__dict__)
        odors_cue = OdorGen(self.list)
        odors_cue.assign_odor()
        self.reward_odor, self.non_reward_odor, self.control_odor = odors_cue.set_rewardodor(index=self.odor_index)
        odors_cue.initiate()
        # odors_cue.odors_DAQ[i]
        print('odor done')

        self.waterR = DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(self.waterline))
        self.waterR.low()
        # self.OdorOnCopy = DAQSimpleDOTask('Dev3/port2/line5')
        # self.OdorOnCopy.low()
        self.lickR = DAQSimpleDITask('Dev2_SELECT/port2/line0')
        print('water done')

        # EVENT CODES
        # video recording start / start trial = 101
        # end trial = 100
        # lick on = 11, lick off = 10

        # reward odor on = 131, off = 130, water on = 51, water off = 50
        # no reward odor on = 141, off = 140
        # c control odor on = 161, off = 160,  water on = 51, water off = 50

        # create excel workbook

        self.wb1 = Workbook()
        self.ws1 = self.wb1.active


        #generate trial type

        # generate trial type
        trialtypes = np.zeros(self.numtrials)
        for i in range(int(self.numtrials/20)):
            train_go = np.zeros(int(round(20 * self.p_go * self.p_reward_go)))  # code 0
            train_nogo = np.ones(int(20 * self.p_no_go))  # code 1
            temp_comb1 = np.concatenate((train_go,train_nogo))
            train_go_omission = np.ones(int(round(20 * self.p_go * (1 - self.p_reward_go)))) * 3  # code 3
            temp_comb1 = np.concatenate((temp_comb1, train_go_omission))
            if self.phase == 'cond':
                train_empty = np.ones(int(20 * self.p_empty* (1-self.p_reward_empty))) * 2  # code 2

                temp_comb1 = np.concatenate((temp_comb1, train_empty))
            elif self.phase == 'deg':
                train_unpredwater = np.ones(int(20 * self.p_empty* self.p_reward_empty)) * 4 # code 4
                temp_comb1 = np.concatenate((temp_comb1, train_unpredwater))
                train_empty = np.ones(int(20 * self.p_empty * (1 - self.p_reward_empty))) * 2  # code 2
                temp_comb1 = np.concatenate((temp_comb1, train_empty))
            elif self.phase == 'close':
                train_unpredwater = np.ones(int(20 * self.p_empty * self.p_reward_empty)) * 7  # code 4
                temp_comb1 = np.concatenate((temp_comb1, train_unpredwater))
                train_empty = np.ones(int(20 * self.p_empty * (1 - self.p_reward_empty))) * 2  # code 2
                temp_comb1 = np.concatenate((temp_comb1, train_empty))
            elif self.phase == 'far':
                train_unpredwater = np.ones(int(20 * self.p_empty * self.p_reward_empty)) * 8  # code 4
                temp_comb1 = np.concatenate((temp_comb1, train_unpredwater))
                train_empty = np.ones(int(20 * self.p_empty * (1 - self.p_reward_empty))) * 2  # code 2
                temp_comb1 = np.concatenate((temp_comb1, train_empty))
            elif self.phase == 'C_control':
                train_ccontrol = np.ones(int(20 * self.p_empty * self.p_reward_empty)) * 5  # code 4
                temp_comb1 = np.concatenate((temp_comb1, train_ccontrol))
                train_ccontrol_nogo = np.ones(int(20 * self.p_empty * (1 - self.p_reward_empty))) * 6  # code 2
                print(len(train_ccontrol_nogo))
                print(len(train_ccontrol))
                temp_comb1 = np.concatenate((temp_comb1, train_ccontrol_nogo))

            temp_comb2 = temp_comb1.copy()
            random.shuffle(temp_comb2)
            trialtypes[20*i:20*(i+1)] = temp_comb2
        self.trialstype = trialtypes
        print('================== Trial Types =================')
        print(trialtypes)
        for t in range(0, self.numtrials):
            print('================================================')
            print('trial number: ', t)

            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=101) #trial start
            d = self.ws1.cell(row=self.ws1.max_row, column=3, value= 'trial{}'.format(int(trialtypes[t]))) #trial type
            # code: 0--go w rward; 1--no go; 2--empty; 3--go w/o reward; 4 --unpred water; 5 -- c control odor； 6 -- c control odor w/o
            if int(trialtypes[t]) in [7,8]:

                #           main training program
                self.run_trial_type(int(trialtypes[t]))


            else:
                self.check_licking_1spout(self.duration_rec_on_before)
                #           main training program
                self.run_trial_type(int(trialtypes[t]))

                self.check_licking_1spout(self.duration_rec_on_after)
            # self.settings.save_video.update_value(False):
            self.check_licking_1spout(self.duration_ITI[t])
            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=100) #end
            self.wb1.save(self.filename)
        odors_cue.initiate()
        odors_cue.close()
        self.waterR.low()
        self.waterR.close()
        print('FINISHED ASSOCIATION TRAINING')

    def check_licking_1spout(self, interval,check_action=False):
        checkperiod = 0.005
        timeout = time.time() + interval
        reward_on = True # this is used for Pav or Operant
        right_lick_last = 0
        count = 0
        while time.time() < timeout:
            right_lick = self.lickR.read()
            if right_lick != right_lick_last:
                if right_lick:
                    print('Lick')
                    d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)
                # self.save_training()
                    if check_action:
                        count += 1
                else:
                    d = self.ws1.cell(row=(self.ws1.max_row+1), column=1, value=time.clock())
                    d = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)
            else:
                pass
            right_lick_last = right_lick
            time.sleep(checkperiod)
        if check_action and count >= self.licknum:

            print('licking activate reward')
        elif check_action and count < self.licknum:
            print('not enough licking')
            reward_on = False
        return reward_on

    def run_trial_type(self, types):
        odor_on = False
        reward_on = False
        is_control = False
        if types == 0:
            print('go trial ' + str(int(self.counter[types])))
            odor_on = True # contingency odor comes
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action)
            reward_on = self.check_licking_1spout(self.duration_action_window, self.operant)  ### can be a problem
            self.run_reward_module(reward_on, w_code)

        elif types == 1:
            print('no go trial ' + str(int(self.counter[types])))

            odor_on = True
            is_go = False
            r_code = [141, 140]
            w_code = [61, 60]
            self.run_odor_module(odor_on,is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action+self.duration_action_window)
            reward_on = False
            self.run_reward_module(reward_on, w_code)
        elif types == 5:
            print('control C trial ' + str(int(self.counter[types])))

            odor_on = True
            is_go = False
            is_control = True
            r_code = [161, 160]
            w_code = [51, 50]
            self.run_odor_module(odor_on,is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action+self.duration_action_window)
            reward_on = self.check_licking_1spout(self.duration_action_window, self.operant)
            self.run_reward_module(reward_on, w_code)

        elif types == 2:
            print('empty trial ' + str(int(self.counter[types])))
            # odor false: control odor comes
            w_code = [71, 70]# not used
           # self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_on)
            self.check_licking_1spout(self.duration_odor_to_action+self.duration_action_window)
            reward_on = False
            self.run_reward_module(reward_on, w_code)

        elif types == 3:
            print('go omission trial ' + str(int(self.counter[types])))
            odor_on = True # contingency odor comes
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action + self.duration_action_window)
            self.run_reward_module(reward_on, w_code)

        elif types == 6:
            print('c odor omission ' + str(int(self.counter[types])))
            odor_on = True # contingency odor comes
            is_go = False
            is_control = True
            r_code = [161, 160]
            w_code = [51, 50]
            self.run_odor_module(odor_on, is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action + self.duration_action_window)
            self.run_reward_module(reward_on, w_code)

        elif types == 4:
            print('unpredicted water trial ' + str(int(self.counter[types])))
            w_code = [51, 50]
            self.check_licking_1spout(self.duration_odor_on)
            self.check_licking_1spout(self.duration_odor_to_action)
            reward_on = self.check_licking_1spout(self.duration_action_window, self.operant)
            self.run_reward_module(reward_on, w_code)

        elif types == 7:
            print('close unpredicted water trial ' + str(int(self.counter[types])))
            w_code = [51, 50]
            reward_on = self.check_licking_1spout(0, self.operant)
            self.run_reward_module(reward_on, w_code)
            self.check_licking_1spout(self.duration_odor_on+2)
            self.check_licking_1spout(self.duration_action_window)

        elif types == 8:
            print('far unpredicted water trial ' + str(int(self.counter[types])))
            w_code = [51, 50]
            self.check_licking_1spout(self.duration_odor_on)
            self.check_licking_1spout(self.duration_odor_to_action+2)
            reward_on = self.check_licking_1spout(self.duration_action_window, self.operant)
            self.run_reward_module(reward_on, w_code)
        self.counter[types] += 1

        self.wb1.save(self.filename)

    def run_odor_module(self, odor_on, is_go, is_control, r_code):
        if odor_on:
            print('opening odor port')
            if is_go and not is_control:
                self.reward_odor.high()
            elif not is_go and not is_control:
                self.non_reward_odor.high()
            elif is_control:
                self.control_odor.high()
            # self.OdorOnCopy.high()  # ？？？
            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=r_code[0])
            # self.save_training()
            self.check_licking_1spout(self.duration_odor_on)
            print('closing odor port')
            if is_go and not is_control:
                self.reward_odor.low()
            elif not is_go and not is_control:
                self.non_reward_odor.low()
            elif is_control:
                self.control_odor.low()

            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=r_code[1])
        else:
            self.check_licking_1spout(self.duration_odor_on)

    def run_reward_module(self, reward_on, w_code):
        if reward_on:
            # modify! give water if licks three times within 1 s
            print('opening water valve')
            self.waterR.high()
            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=w_code[0])
            self.check_licking_1spout(self.duration_water_large)  # this parameter hasn't een defined
            print('closing water valve')
            self.waterR.low()
            d = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d = self.ws1.cell(row=self.ws1.max_row, column=2, value=w_code[1])
        else:
            self.check_licking_1spout(self.duration_water_large)






#%%
class Mouse_data:
    def __init__(self,mouse_id,filedir):
        self.mouse_id = mouse_id
        self.filedir = filedir
        self.filename = ''
        self.selected_filename = ''
        self.all_days = []
        
        self.training_type = []
        self.df_trials = {}
        self.trialtypes = []
        self.df_trials_iscorrect = {}
        self.df_trials_lick = {}
        self.df_eventcode = {}
        self.p_hit = {}
        self.p_correj = {}
        self.licking_actionwindow = {}
        self.licking_latency = {}
        self.licking_baselicking = {}
        self.stats = {}
        self.event_data = ''
        self.odor_bef = 4.0
        self.odor_on = 1.0
        self.delay = 2.5
        self.rew_after = 4.0
        

    def read_filename(self):
        filedir = self.filedir +'/{}'.format(self.mouse_id)
        filename = []
        for dirpath, dirnames, files in os.walk(filedir): # can walk through all levels down
        #     print(f'Found directory: {dirpath}')
            for f_name in files:
                if f_name.endswith('.xlsx'):
                    filename.append(dirpath+'/'+f_name)
                    print(f_name)
        print('---------------------------------------------')    
        print('The files have been loaded from the following paths')
        print(filename[-45:])
        self.filename = filename
        
    
    def select_dates(self):
        pass
    def delete_date(self, dates):
        for date in dates:
            self.all_days.remove(date)
        
        return self.all_days
        
    def create_eventcode(self, original = True): #{'date':df of eventcode} format from original CSV
        date_list = []
        df = {}
        training_type = []
        
        for file in self.filename:
            date = re.search(r"(\d{4}-\d{1,2}-\d{1,2}-\d{1,2}-\d{1,2})",file).group(0) # extract date: must be like format 2020-02-10
            date_list.append(date) # create a list of emperiment date
            
            train_type = os.path.split(file)[-1][16:-5]
            training_type.append(train_type) ###
                
            
            data = pd.read_excel(file,header = 0 if original else None) #read orginal csv data
            data.columns = ['Time','Event','Type'] # rename columns
            
            df.update({date:data}) # create the dict of key: date and value: data dataframe
        self.df_eventcode = df #individual mouse event code data
        date_format = '%Y-%m-%d-%h-%s'
        index = np.argsort(date_list)
        self.all_days = [date_list[i] for i in index]
        self.training_type = [training_type[i] for i in index]
        
        print('---------------------------------------------')
        print('{0} has data from these days: {1}'.format(self.mouse_id,zip(self.all_days,self.training_type)))

    def create_trials(self): #{'date: df of trials}
        for index , date in enumerate(self.all_days):
            value = self.df_eventcode[date]
        
            new_df = self.generate_trials_dataframe(index,value)
            self.df_trials[date] = new_df
            print('{} done!'.format(date))
     
    
    def generate_trials_dataframe(self,index,ori_df):
        
        lick, trialtype, go_odor, nogo_odor, control_odor, water_on, water_off, trial_end = self.seperate_events(index,ori_df)
        d = {'trialtype' : trialtype, 'go_odor': go_odor, 'nogo_odor': nogo_odor, 'control_odor':control_odor,
             'water_on':water_on,'water_off':water_off,'licking':lick,
             'trial_end':trial_end}
        df_trial = pd.DataFrame(data = d)
        return df_trial
    
    
    def seperate_events(self,index_day,df):
        
        start_trials = 0
        lick          = []
        trialtype     = []
        go_odor       = []
        nogo_odor     = []
        control_odor  = []
        water_on      = []
        water_off     = []
        trial_end     = []
        print(index_day)
        
        for index, row in df.iterrows():
            if row['Event'] == 101:   # trial start
                start_trials = row['Time']                
                temp_licks        = []
                temp_go_odor_on      = np.nan
                temp_go_odor_off      = np.nan
                temp_nogo_odor_on    = np.nan
                temp_nogo_odor_off    = np.nan
                temp_control_odor_on = np.nan
                temp_control_odor_off = np.nan
                
                temp_water_on     = np.nan
                temp_water_off    = np.nan
                temp_trial_end    = np.nan
                
                if row['Type'] == 'trial0':
                    trialtype.append('go')
                elif row['Type'] == 'trial1':
                    trialtype.append('no_go')
                elif row['Type'] == 'trial2':
                    trialtype.append('background')
                elif row['Type'] == 'trial3':
                    trialtype.append('go_omit')
                elif row['Type'] == 'trial4':
                    trialtype.append('unpred_water')
                elif row['Type'] == 'trial5':
                    trialtype.append('c_reward')
                elif row['Type'] == 'trial6':
                    trialtype.append('c_omit')
                elif row['Type'] == 'trial7':
                    trialtype.append('close_unpred_water')
                elif row['Type'] == 'trial8':
                    trialtype.append('far_unpred_water')

            elif row['Event'] == 11:
                lick_time = row['Time'] - start_trials
                temp_licks.append(lick_time)

            elif row['Event'] == 131: # go odor on
                temp_go_odor_on=row['Time']- start_trials
        
            elif row['Event'] == 130:
                temp_go_odor_off=row['Time']- start_trials

            elif row['Event'] == 141: # no go odor on
                temp_nogo_odor_on=row['Time']- start_trials
        
            elif row['Event'] == 140:
                temp_nogo_odor_off=row['Time']- start_trials
            elif row['Event'] == 161: # no go odor on
                temp_control_odor_on=row['Time']- start_trials
        
            elif row['Event'] == 160:
                temp_control_odor_off=row['Time']- start_trials
 
            elif row['Event'] == 51: # water on
                temp_water_on=row['Time']- start_trials
        
            elif row['Event'] == 50:
                temp_water_off=row['Time']- start_trials

            elif row['Event'] == 100: #trial end
                temp_trial_end=row['Time']- start_trials
    
                lick.append(temp_licks)
                go_odor.append([temp_go_odor_on,temp_go_odor_off])
                
                nogo_odor.append([temp_nogo_odor_on,temp_nogo_odor_off])
                control_odor.append([temp_control_odor_on,temp_control_odor_off])
                water_on.append(temp_water_on)
                water_off.append(temp_water_off)
                trial_end.append(temp_trial_end)
        
    
        return  lick, trialtype, go_odor, nogo_odor, control_odor, water_on, water_off, trial_end
    
    def create_trial_iscorrect(self): # create dataframe with trial number, correct or rewarded or not only for conditioning period
        for index , date in enumerate(self.all_days):    
            value = self.df_trials[date]
            new_df = self.eval_trials_correct(value)
            new_df.insert(0,'trialtype',value['trialtype'])
            self.df_trials_iscorrect[date] = new_df
            print('create_trial_iscorrect done!')
            
    def eval_trials_correct(self, df):
        
        is_correct = []
        is_rewarded = []
        for index, row in df.iterrows():
            if row['trialtype'] == 'go':
                is_rewarded.append(1)
                if any(x > row['go_odor'][0] and x < row['go_odor'][1]+self.delay for x in row['licking']):
                    is_correct.append(1)
                else:
                    is_correct.append(0)
                
            elif row['trialtype'] == 'no_go':
                is_rewarded.append(0)
                if any(x > row['nogo_odor'][0] and x < row['nogo_odor'][1]+self.delay for x in row['licking']):
                    is_correct.append(0)
                else:
                    is_correct.append(1)
            
            elif row['trialtype'] == 'c_reward':
                is_rewarded.append(1)
                if any(x > row['control_odor'][0] and x < row['water_on'] for x in row['licking']):
                    is_correct.append(1)
                else:
                    is_correct.append(0)
            elif row['trialtype'] == 'c_omit':
                is_rewarded.append(0)
                if any(x > row['control_odor'][0] and x < row['control_odor'][1]+2*self.delay for x in row['licking']):
                    is_correct.append(1)
                else:
                    is_correct.append(0)
            
            elif row['trialtype'] == 'background':
                is_rewarded.append(0)
                if any(x > 0 and x < row['trial_end'] for x in row['licking']):
                    is_correct.append(0)
                else:
                    is_correct.append(1)
                
            elif row['trialtype'] == 'go_omit':
                is_rewarded.append(0)
                if any(x > row['go_odor'][0] and x < row['go_odor'][1]+self.delay for x in row['licking']):
                    is_correct.append(1)
                else:
                    is_correct.append(0)
            elif row['trialtype'] in ['unpred_water','close_unpred_water','far_unpred_water']:
                is_rewarded.append(1)
                is_correct.append(np.nan)
        d = {'is_Correct':is_correct,'is_Rewarded':is_rewarded}
        new_df = pd.DataFrame(d)
        return new_df
    
    def create_trial_lick(self):
        for index , date in enumerate(self.all_days):
            value = self.df_trials[date]            
            new_df = self.lick_stats(value)               
            new_df.insert(0,'trialtype',value['trialtype'])
            self.df_trials_lick[date] = new_df
            print('lick stats done!')        
            
    def lick_stats(self, df):
        
        lick_num = []
        lick_rate = []
        lick_latent_odor = []
        lick_latent_rew = []
        lick_duration = []
        lick_rate_anti = []
        lick_rate_aftr = []
        tol_interval = self.odor_on + self.delay + self.rew_after
        anti_window = self.odor_on + self.delay
        lick_anti_list = []
        lick_aftr_list = []
        for index, row in df.iterrows():
            if row['trialtype'] in ['go','go_omit'] :
                
                lick_valid = [x for x in row['licking'] if x > row['go_odor'][0] and x < row['go_odor'][0]+tol_interval ] # valid licking: after odor on and 5.5 s after odor off
                #lickingrate for anticipitory period and after water period 3.5 respectively
                anti = [i for i in row['licking']  if i> row['go_odor'][0] and i< row['go_odor'][1]+self.delay] #anticipitory
                aftr = [i for i in row['licking']  if i> row['water_on'] and i< row['water_off']+self.rew_after] #after water
                rate_anti = len(anti)/anti_window                
                rate_aftr = len(aftr)/self.rew_after             
                #num of licking
                num = len(lick_valid)
                if num != 0:
                    latency_odor = min(lick_valid)-row['go_odor'][0]# first licking after odor delivery on
                else:
                    latency_odor = tol_interval
                if row['trialtype'] == 'go':
                    if len(aftr) != 0:
                        latency_rew = min(aftr)-row['water_on']# first licking after odor delivery on
                    else:
                        latency_rew = self.rew_after
                else:
                    latency_rew = np.nan
                try:    
                    duration = max(anti)-min(anti) # anticipitory licking duration after odor presentation
                except:
                    duration = np.nan
                
            elif row['trialtype'] == 'no_go':
                lick_valid = [x for x in row['licking'] if x > row['nogo_odor'][0] and x < row['nogo_odor'][0]+tol_interval ] # valid licking: after odor on and 5.5 s after odor off
                #inter-licking interval for anticipitory period and after water period
                anti = [i for i in row['licking']  if i> row['nogo_odor'][0] and i< row['nogo_odor'][1]+self.delay] #anticipitory
                aftr = []               
                rate_anti = len(anti)/anti_window           
                rate_aftr = np.nan                
                #num of licking
                num = len(lick_valid)
                if num != 0:
                    latency_odor = min(lick_valid)-row['nogo_odor'][0]# first licking after odor delivery on
                else:
                    latency_odor = tol_interval
                latency_rew = np.nan
                try:    
                    duration = max(anti)-min(anti) # anticipitory licking duration after odor presentation
                except:
                    duration = np.nan
            elif row['trialtype'] in ['c_reward','c_omit'] :
                
                lick_valid = [x for x in row['licking'] if x > row['control_odor'][0] and x < row['control_odor'][0]+tol_interval+self.delay ] # valid licking: after odor on and 5.5 s after odor off
                #lickingrate for anticipitory period and after water period 3.5 respectively
                anti = [i for i in row['licking']  if i> row['control_odor'][0] and i< row['water_on']] #anticipitory
                aftr = [i for i in row['licking']  if i> row['water_on'] and i< row['water_off']+self.rew_after] #after water
                rate_anti = len(anti)/(row['water_on']- row['control_odor'][0])           
                rate_aftr = len(aftr)/self.rew_after             
                #num of licking
                num = len(lick_valid)
                if num != 0:
                    latency_odor = min(lick_valid)-row['control_odor'][0]# first licking after odor delivery on
                else:
                    latency_odor = row['water_on']- row['control_odor'][0]
                if row['trialtype'] == 'c_reward':
                    if len(aftr) != 0:
                        latency_rew = min(aftr)-row['water_on']# first licking after odor delivery on
                    else:
                        latency_rew = self.rew_after
                else:
                    latency_rew = np.nan
                try:    
                    duration = max(anti)-min(anti) # anticipitory licking duration after odor presentation
                except:
                    duration = np.nan
                
            elif row['trialtype'] == 'background':
                lick_valid = [x for x in row['licking'] if x > 0 and x < row['trial_end']] # valid licking: after odor on and 5.5 s after odor off
                intvl = row['trial_end']
                anti = lick_valid
                aftr = []
                rate_anti = len(anti)/intvl    
                rate_aftr = np.nan                
                num = len(lick_valid)
                latency_odor = np.nan
                latency_rew = np.nan
                duration = np.nan # licking duration after water delivery
                
            elif row['trialtype'] in ['unpred_water','close_unpred_water','far_unpred_water']:
                lick_valid = [x for x in row['licking'] if x > 0 and x < row['trial_end'] ] # valid licking: after odor on and 5.5 s after odor off                
                #inter-licking interval for anticipitory period and after water period
                anti = [i for i in row['licking']  if i> 0 and i< row['water_on']] #anticipitory
                
                aftr = [i for i in row['licking']  if i> row['water_on'] and i < min((row['water_off']+self.rew_after),row['trial_end'])]               
                rate_anti = len(anti)/(row['water_on'])         
                rate_aftr = len(aftr)/min(self.rew_after,row['trial_end']-row['water_off'])   
                #num of licking
                num = len(lick_valid)
                if len(aftr) != 0:
                    latency_rew = min(aftr)-row['water_on']# first licking after unpredicted water
                else:
                    latency_rew = min(self.rew_after,row['trial_end']-row['water_off'])
                latency_odor = np.nan
                try:    
                    duration = max(aftr)-min(aftr) # anticipitory licking duration after odor presentation
                except:
                    duration = np.nan
            
              
            lick_num.append(num)
            lick_rate.append(num/tol_interval)
            lick_latent_odor.append(latency_odor)
            lick_latent_rew.append(latency_rew)
            lick_duration.append(duration)
            lick_rate_anti.append(rate_anti)
            lick_rate_aftr.append(rate_aftr)
            lick_anti_list.append(anti)
            lick_aftr_list.append(aftr)
                
        d = {'lick_num_whole_trial':lick_num,
             'lick_rate_whole_trial':lick_rate,
             'latency_to_odor':lick_latent_odor,
             'latency_to_rew':lick_latent_rew,
             'anti_duration':lick_duration,
             'rate_antici':lick_rate_anti,
             'rate_after':lick_rate_aftr,
             'anti_lick':lick_anti_list,
             'aftr_lick':lick_aftr_list
             }
        new_df = pd.DataFrame(d)
        return new_df
    

        
    
def save_to_excel(dict_df,path,filename):
    try:
        os.makedirs(path) # create the path first
    except FileExistsError:
        print('the path exist.')
    filename = path +'/{}.xlsx'.format(filename)

    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(filename, engine='xlsxwriter')
    
    # Write each dataframe to a different worksheet. you could write different string like above if you want
    for key, value in dict_df.items():
        value.to_excel(writer, sheet_name= key)
        
    
    # Close the Pandas Excel writer and output the Excel file.
    writer.save()
    print('save to excel done!')
    
def pickle_dict(df,path,filename):
    try:
        os.makedirs(path) # create the path first
    except FileExistsError:
        print('the path exist.')
    filename = path +'/{}.pickle'.format(filename)
    with open(filename, 'wb') as handle:
        pickle.dump(df, handle, protocol=4)
    print('save to pickle done!')


def load_pickleddata(filename):
    
    with open(filename, 'rb') as handle:
        df = pickle.load(handle)
    return df
                   

def event_plot(df,save_dir,mouse_id = '', exp_date = '', filename = 'test',save = False, width = 3.5, figuresize = [12,20],gocolor = '#EBA0B1',nogocolor = '#F9CD69', rewardcolor = '#3083D1',lickcolor = 'grey'):
    lineoffsets2 = 1
    linelengths2 = 1
    
    
    # create a horizontal plot
    figure = plt.figure()

    for i, row in df.iterrows():
        
        
        plt.hlines(i, row['go_odor'][0], row['go_odor'][1],color = gocolor,alpha = 1,
                   linewidth = width,label = 'go odor' if i ==0 else '')
        plt.hlines(i, row['water_on'], row['water_off'],color = rewardcolor,
                   linewidth = width,label = 'water' if i ==0 else '')
        plt.hlines(i, row['nogo_odor'][0], row['nogo_odor'][1],color = nogocolor,alpha = 1,
                   linewidth = width,label = 'no go odor' if i ==0 else '')
        plt.hlines(i, 0, row['trial_end'],color = 'grey',alpha = 0.2,
                   linewidth = width)
        plt.hlines(i, row['control_odor'][0], row['control_odor'][1],alpha = 1,
                   linewidth = width,label = 'control odor' if i ==0 else '')
        
    plt.eventplot(df.licking, colors=lickcolor, lineoffsets=lineoffsets2,linelengths=linelengths2,alpha = 0.5, label = 'licking')
    
    ax = plt.subplot(111)    
    ax.spines["top"].set_visible(False)    
    ax.spines["right"].set_visible(False)    
    draw_loc = ax.get_xlim()[1]
    draw_loc2 = ax.get_ylim()[1]
    # create is correct
    try:
        
        for i in range(len(df.index)):
            #try:
            appear1 = 0
            appear2 = 0
            if df['trialtype'][i] in ['go','go_omit','c_reward','c_omit']:
                position = 1
            elif df['trialtype'][i] in ['no_go','background']:
                position = 2.3
            if df['is_Correct'][i]:
                plt.hlines(i, draw_loc+position, draw_loc+position+0.5,
                           color = '#A8DEA3',alpha = 1,linewidth = width,label = 'correct' if appear1 ==0 else '' ) #green
                appear1 = 1
            else:
                plt.hlines(i, draw_loc+position, draw_loc+position+0.5,
                           color = '#F9634B',alpha = 1,linewidth = width ,label = 'wrong' if appear2 ==0 else '') # red
                appear2 = 1
    
            
            
            #except:
            #    print("--------is_Correct data hasn't been merged ---------")
        #plt.xticks(fontsize=14)    
        ax.set_xlim([ax.get_xlim()[0],draw_loc+4])
    except:
        pass
    
    ax.get_xaxis().tick_bottom() 
    
    ax.get_yaxis().tick_left()
    plt.tick_params(axis="both", which="both", bottom=False, left = False, top = False, right = False,
                    labelbottom="on", labelleft="on",labelsize = 14)   
    plt.ylabel('Trials',fontsize = 18)
    plt.xlabel('Time(s)',fontsize = 18)
    ax.set_ylim(bottom=-1,ymax = len(df.index))
    ax.set_xlim(left=-0.2)
    
    plt.text(draw_loc +0.8, draw_loc2-3, 'hit', fontsize=14)
    plt.text(draw_loc +1.6, draw_loc2-3, 'rejection', fontsize=14)
    
    handles, labels = plt.gca().get_legend_handles_labels()
    by_label = OrderedDict(zip(labels, handles))
    plt.legend(by_label.values(), by_label.keys(),prop={'size': 14},loc='upper center', bbox_to_anchor=(0.5, 1.07),
          frameon=False,fancybox=False, shadow=False, ncol=3)
    datetime.now().strftime("%Y-%m-%d-%H-%M")
    plt.suptitle('{} eventplot on {}'.format(mouse_id, exp_date),fontsize = 20, y = 0.96)
    if save:
        try:
            savepath = "{1}/{0}/{2}".format(mouse_id,save_dir,date.today())
            os.makedirs(savepath)
        except:
            pass
        figure.set_size_inches(figuresize[0], figuresize[1])
    
        plt.savefig("{0}/{1}_{2}.png".format(savepath,exp_date,filename), bbox_inches="tight", dpi = 100)

        
        #    print('error while saving')
    
    plt.show()
    return figure
    
        


#%% main code
if __name__ == '__main__':
    # running session
    session = SelinaTraining()
    print('Session starts')
    session.run()
    # plotting session figure
    is_save = True
    is_original = True # when use clean data, change it to False
    
    #********************
    load_path = session.events_path
    
    # load file
    mouse_names = session.mouse
    

    mouse = Mouse_data(mouse_name, filedir = load_path)
    mouse.read_filename()
    #parse data
    mouse.create_eventcode(original = is_original)
    mouse.create_trials()
    mouse.create_trial_iscorrect()    
    mouse.create_trial_lick()
    mouse_trials = mouse.df_trials
    mouse_iscorrect = mouse.df_trials_iscorrect
    # choose a date
    all_days = mouse.all_days
    print('-----------------------------------------------------------')
    print('there are ', len(all_days),'days, condition days is ', len([x for x in mouse.training_type if x == 'cond']))
    for index in range(len(all_days)):
        print('you are looking at day ', all_days[index],'training type is ',mouse.training_type[index] )
        day = all_days[index] 
        
        save_dir = os.path.join(path,'figures')
        # concatenate data
        merge_trials_iscorrect = mouse_trials[day].copy() # need to be hard copy
        try: # for condition days
            merge_trials_iscorrect['is_Correct'] = mouse_iscorrect[day]['is_Correct']# join df_trials and df_isccorect serial
        except: # degradation days don't have is_correct
            pass
        # plot
        figure = event_plot(merge_trials_iscorrect,mouse_id = mouse.mouse_id,exp_date = day,filename = 'all_trials',save = True,save_dir = save_dir)
        
        # only choose go trials for above day
        is_go_trials = merge_trials_iscorrect['trialtype'] == 'go' #or 'go_omit' # index of go trials
        merged_go_trials = merge_trials_iscorrect[is_go_trials] # select out the go trials from the merged dataset
        merged_go_trials.index = range(len(merged_go_trials)) # reindex the index; otherwise the index will be original index like 1,4,14,23,etc
        # plot
        figure = event_plot(merged_go_trials,mouse_id = mouse.mouse_id,exp_date = day,filename = 'go_trials',save = True, save_dir = save_dir,width = 3,figuresize = [12,15])
        
        # only choose no_go trials for above day
        is_nogo_trials = merge_trials_iscorrect['trialtype'] == 'no_go' # index of no_go trials
        merged_nogo_trials = merge_trials_iscorrect[is_nogo_trials] # select out the no_go trials from the merged dataset
        merged_nogo_trials.index = range(len(merged_nogo_trials)) # reindex the index; otherwise the index will be original index like 1,4,14,23,etc
        # plot
        figure = event_plot(merged_nogo_trials,mouse_id = mouse.mouse_id,exp_date = day,filename = 'no_go_trials',save = True, save_dir = save_dir,width = 3,figuresize = [12,15])
        
        # only choose empty trials for above day
        is_background_trials = merge_trials_iscorrect['trialtype'] == 'background' # index of background trials
        merged_background_trials = merge_trials_iscorrect[is_background_trials] # select out the background trials from the merged dataset
        merged_background_trials.index = range(len(merged_background_trials)) # reindex the index; otherwise the index will be original index like 1,4,14,23,etc
        # plot
        figure = event_plot(merged_background_trials,mouse_id = mouse.mouse_id,exp_date = day,filename = 'background_trials',save = True, save_dir = save_dir,width = 3,figuresize = [12,15])
    
    
    




