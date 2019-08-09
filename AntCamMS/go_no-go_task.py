'''
Created on July 17, 2019

@author: Selina Qian
'''
from ScopeFoundry import Measurement
import datetime
import numpy as np
import random
import pickle
import time
from AntCamHW.daq_do.daq_do_dev import DAQSimpleDOTask
from AntCamHW.daq_di.daq_di_dev import DAQSimpleDITask
from openpyxl import Workbook
import os
import logging

class OdorGen(object):
    def __init__(self,odorindex):
        self.odorindex = odorindex
        self.odors_DAQ = []


    def assign_odor(self):
        # initiate all the odor solenoids
        for item in self.odorindex:
            self.odors_DAQ.append(DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(item)))
        print('Odor {} has been properly assigned'.format(self.odorindex))

    def set_rewardodor(self, index: list):
        reward_odor = self.odors_DAQ[index[0]]
        non_reward_odor = self.odors_DAQ[index[1]]
        print('reward odor is odor {}'.format(index))
        return reward_odor,  non_reward_odor

    def initiate(self):
        for odor in self.odors_DAQ:
            odor.low()
        print('Odor initiation: status low')

    def close(self):
        for odor in self.odors_DAQ:
            odor.close()
        print('Connection has been closed')


class SelinaTraining(Measurement):
    def __init__(self):
        self.list = [7, 6]
        self.events_path = "C:/Users/MurthyLab/Desktop/Selina/experiment_data/C16/"+datetime.datetime.now().strftime("%Y-%m-%d")+"/"
        self.events_filename = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M")+'session_1.xlsx'
        self.reward_odor_index = [1, 0] #odor list index change according to mi
        self.operant = True
        self.licknum = 1


        self.numtrials = 100 #need to be 20*
        # pre training
        self.p_go = 0.8
        self.p_no_go = 0.1
        self.p_empty = 0.1

        self.counter = np.zeros(3)

        self.duration_rec_on_before = 2
        self.duration_odor_on = 1
        self.duration_odor_to_action = 0
        self.duration_action_window = 2.5
        self.duration_water_large = 0.1
        self.duration_rec_on_after = 4
        self.duration_ITI = np.random.exponential(2, size=self.numtrials)

        self.waterline = 0
        self.filename = self.events_path + self.events_filename

    def run(self):
        try:
            os.mkdir(self.events_path)
        except OSError:
            print("The directory %s existed" % self.events_path)
        else:
            print("Successfully created the directory %s " % self.events_path)
        logname = self.filename[0:-5] + '.log'
        logging.basicConfig(filename=logname, level=logging.DEBUG)
        logging.info(self.__dict__)
        odors_cue = OdorGen(self.list)
        odors_cue.assign_odor()
        self.reward_odor, self.non_reward_odor = odors_cue.set_rewardodor(index=self.reward_odor_index)
        odors_cue.initiate()
        # odors_cue.odors_DAQ[i]
        print('odor done')

        self.waterR = DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(self.waterline))
        self.waterR.low()
        # self.OdorOnCopy = DAQSimpleDOTask('Dev3/port2/line5')
        # self.OdorOnCopy.low()
        self.lickR = DAQSimpleDITask('Dev2_SELECT/port1/line0')
        print('water done')

        # EVENT CODES
        # video recording start / start trial = 101
        # end trial = 100
        # lick on = 11, lick off = 10

        # contingency reward odor on = 131, off = 130, water on = 51, right water off = 50
        # contingency no reward odor on = 141, off = 140, water on = 61, right water off = 60
        # non-contingency reward odor on = 151, off = 150, water on = 71, right water off = 70
        # non-contingency no reward odor on = 161, off = 160, water on = 81, right water off = 80

        # create excel workbook

        self.wb = Workbook()
        self.ws = self.wb.active


        #generate trial type

        # generate trial type
        trialtypes = np.zeros(self.numtrials)
        for i in range(int(self.numtrials/20)):
            train_go = np.zeros(int(20 * self.p_go))  # code 0
            train_nogo = np.ones(int(20 * self.p_no_go))  # code 1
            temp_comb1 = np.concatenate((train_go,train_nogo))

            train_empty = np.ones(int(20 * self.p_empty)) * 2  # code 2

            temp_comb2 = np.concatenate((temp_comb1, train_empty))
            random.shuffle(temp_comb2)
            trialtypes[20*i:20*(i+1)] = temp_comb2


        self.trialstype = trialtypes
        print('================== Trial Types =================')
        print(trialtypes)

        for t in range(0, self.numtrials):
            print('================================================')

            print('trial number: ', t)


            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=101)
            self.check_licking_1spout(self.duration_rec_on_before)

#           main training program
            self.run_trial_type(int(trialtypes[t]))

            self.check_licking_1spout(self.duration_rec_on_after)

            # self.settings.save_video.update_value(False):
            self.check_licking_1spout(self.duration_ITI[t])
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=100)
            self.wb.save(self.filename)

        odors_cue.initiate()
        odors_cue.close()
        self.waterR.low()
        self.waterR.close()

        print('FINISHED ASSOCIATION TRAINING')




    def check_licking_1spout(self, interval,check_action=False):

        checkperiod = 0.01
        timeout = time.time() + interval
        reward_on = True
        right_lick_last = 0

        count = 0
        while time.time() < timeout:
            right_lick = self.lickR.read()

            if right_lick != right_lick_last:
                if right_lick:

                    print('Lick')
                    d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
                    d = self.ws.cell(row=self.ws.max_row, column=2, value=11)
                # self.save_training()
                    if check_action:
                        count += 1
                else:

                    d = self.ws.cell(row=(self.ws.max_row+1), column=1, value=time.clock())
                    d = self.ws.cell(row=self.ws.max_row, column=2, value=10)
            else:
                pass
            right_lick_last = right_lick
            time.sleep(checkperiod)
        if check_action and count >= self.licknum:

            print('licking activate reward')
        elif check_action and count < self.licknum:
            print('not enough licking')
            reward_on = False
        return reward_on

    def run_trial_type(self,types):
        odor_on = False
        reward_on = False
        if types == 0:
            print('go trial ' + str(int(self.counter[types])))

            odor_on = True # contingency odor comes
            is_go = True

            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, is_go, r_code)
            self.check_licking_1spout(self.duration_odor_to_action)
            reward_on = self.check_licking_1spout(self.duration_action_window, self.operant)  ### can be a problem
            self.run_reward_module(reward_on, w_code)

        elif types == 1:
            print('no go trial ' + str(int(self.counter[types])))

            odor_on = True
            is_go = False
            r_code = [141, 140]
            w_code = [61, 60]
            self.run_odor_module(odor_on,is_go, r_code)
            self.check_licking_1spout(self.duration_odor_to_action+self.duration_action_window)  ### can be a problem
            rewar_on = False
            self.run_reward_module(reward_on, w_code)
        elif types == 2:
            print('empty trial ' + str(int(self.counter[types])))
            # odor false: control odor comes

            r_code = [151, 150]
            w_code = [71, 70]
           # self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_on)
            self.check_licking_1spout(self.duration_odor_to_action+self.duration_action_window)  ### can be a problem
            reward_on = False
            self.run_reward_module(reward_on, w_code)

        self.counter[types] += 1

        self.wb.save(self.filename)

    def run_odor_module(self,odor_on, is_go, r_code):
        if odor_on:
            print('opening odor port')
            if is_go:
                self.reward_odor.high()
            else:
                self.non_reward_odor.high()
            # self.OdorOnCopy.high()  # ？？？
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[0])
            # self.save_training()

            self.check_licking_1spout(self.duration_odor_on)

            print('closing odor port')
            if is_go:
                self.reward_odor.low()
            else:
                self.non_reward_odor.low()

            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[1])

        else:

            self.check_licking_1spout(self.duration_odor_on)




    def run_reward_module(self ,reward_on, w_code):
        if reward_on:

            # modify! give water if licks three times within 1 s

            print('opening water valve')
            self.waterR.high()
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[0])

            self.check_licking_1spout(self.duration_water_large)  # this parameter hasn't een defined

            print('closing water valve')
            self.waterR.low()
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[1])


        else:

            self.check_licking_1spout(self.duration_water_large)


#     def camera_action(self):
#         '''
#         format the image properly
#         '''
#         try:
#             wide_image = self.wide_cam.read()
#             wide_image_data = self.wide_cam.to_numpy(wide_image)
#             self.wide_disp_queue.put(wide_image_data)
#
#             if self.settings.save_video.value() and self.settings.in_trial.value() :
#                 self.recorder.save_frame(self.settings.filename.value(),wide_image)
#
#         except Exception as ex:
#             print('Error : %s' % ex)

    def save_training(self):
        with open(self.filename, 'wb') as output:
            pickle.dump(self, output, pickle.HIGHEST_PROTOCOL)




test = SelinaTraining()
print('start')
test.run()


