'''
Created on July 17, 2019
the code for running two setup in parallel
@author: Selina Qian
'''
from ScopeFoundry import Measurement
import datetime
import numpy as np
import random
import pickle
import time
from AntCamHW.daq_do.daq_do_dev import DAQSimpleDOTask
from AntCamHW.daq_di.daq_di_dev import DAQSimpleDITask
from openpyxl import Workbook
import os
import logging

class OdorGen(object):
    def __init__(self,odorindex):
        self.odorindex = odorindex
        self.odors_DAQ = []


    def assign_odor(self):
        # initiate all the odor solenoids
        for item in self.odorindex:
            self.odors_DAQ.append(DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(item)))
        print('Odor {} has been properly assigned'.format(self.odorindex))

    def set_rewardodor(self, index: list):
        reward_odor = self.odors_DAQ[index[0]]
        non_reward_odor = self.odors_DAQ[index[1]]
        if len(self.odorindex) == 3:
            control_odor = self.odors_DAQ[index[2]]

            print('reward odor, unrewarded odor and control odor loaded')
            return reward_odor, non_reward_odor, control_odor
        else:
            print('reward odor, unrewarded odor loaded')
            return reward_odor,  non_reward_odor

    def initiate(self):
        for odor in self.odors_DAQ:
            odor.low()
        print('Odor initiation: status low')

    def close(self):
        for odor in self.odors_DAQ:
            odor.close()
        print('Connection has been closed')


class SelinaTraining(Measurement):
    def __init__(self):
        # please change this according to mouse
        # two mice together
        self.mouse = ['C35','C36']  #OT-GC-2

        self.phase = 'cond' #'cond', 'deg','C_control','close','far'
        self.condition = 'Pav'
        self.numtrials = 20

        self.list = [7, 6, 5]
        # two file saving path
        self.events_path_0 = "C:/Users/MurthyLab/Desktop/Selina_lab_computer_data/experiment_data_2021_7_{0}/{1}/".format(
            self.condition,self.mouse[0])
        self.events_path_1 = "C:/Users/MurthyLab/Desktop/Selina_lab_computer_data/experiment_data_2021_7_{0}/{1}/".format(
            self.condition, self.mouse[1])
        self.events_filename = datetime.datetime.now().strftime("%Y-%m-%d-%H-%M")+'{}.xlsx'.format(self.phase)

        self.filename_0 = self.events_path_0 + self.events_filename
        self.filename_1 = self.events_path_1 + self.events_filename

        self.odor_index = [1,0,2] #odor list index position 0 is reward, 1 is unrewarded, 2 is control c odor; so here, 7 is unrewarded, 6 is reward and 5 is c odor
        if self.condition == 'Operant':
            self.operant = True
            self.licknum = 1
        else:
            self.operant = False
            self.licknum = 0


        # pre training
        self.p_go = 0.4#0.5
        self.p_reward_go = 0.75
        self.p_no_go = 0.2#0.2
        self.p_empty = 0.4#0.3
        if self.phase == 'cond':
            self.p_reward_empty = 0  # 0.3
        else: # 'deg','C-control','close','far'
            self.p_reward_empty = 0.75


        self.counter = np.zeros(9)

        self.duration_rec_on_before = 1
        self.duration_odor_on = 1
        self.duration_odor_to_action = 0
        self.duration_action_window = 2.5
        self.duration_water_large = 0.08 #0.1
        self.duration_rec_on_after = 1
        self.duration_ITI = np.random.exponential(1, size=self.numtrials)

        self.waterline_0 = 0
        self.waterline_1 = 1


    def run(self):
        try:
            os.makedirs(self.events_path_0)
            os.makedirs(self.events_path_1)
        except OSError:
            print("The directory %s existed" % self.events_path_0)
        else:
            print("Successfully created the directory %s " % self.events_path_0)
        logname_0 = self.filename_0[0:-5] + '.log'
        logname_1 = self.filename_1[0:-5] + '.log'
        logging.basicConfig(filename=logname_0, level=logging.DEBUG)
        logging.info(self.__dict__)
        logging.basicConfig(filename=logname_1, level=logging.DEBUG)
        logging.info(self.__dict__)
        self.lick_0 = DAQSimpleDITask('Dev2_SELECT/port2/line0')
        self.lick_1 = DAQSimpleDITask('Dev2_SELECT/port2/line1')
        print('water done')
        odors_cue = OdorGen(self.list)
        odors_cue.assign_odor()
        self.reward_odor, self.non_reward_odor, self.control_odor = odors_cue.set_rewardodor(index=self.odor_index)
        odors_cue.initiate()
        # odors_cue.odors_DAQ[i]
        print('odor done')

        self.water_0 = DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(self.waterline_0))
        self.water_0.low()
        self.water_1 = DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(self.waterline_1))
        self.water_1.low()
        # self.OdorOnCopy = DAQSimpleDOTask('Dev3/port2/line5')
        # self.OdorOnCopy.low()


        # EVENT CODES
        # video recording start / start trial = 101
        # end trial = 100
        # lick on = 11, lick off = 10

        # reward odor on = 131, off = 130, water on = 51, water off = 50
        # no reward odor on = 141, off = 140
        # c control odor on = 161, off = 160,  water on = 51, water off = 50

        # create two excel workbooks
        self.wb0 = Workbook()
        self.ws0 = self.wb0.active
        self.wb1 = Workbook()
        self.ws1 = self.wb1.active


        #generate trial type

        # generate trial type
        trialtypes = np.zeros(self.numtrials)
        for i in range(int(self.numtrials/20)):
            train_go = np.zeros(int(round(20 * self.p_go * self.p_reward_go)))  # code 0
            train_nogo = np.ones(int(20 * self.p_no_go))  # code 1
            temp_comb1 = np.concatenate((train_go,train_nogo))
            train_go_omission = np.ones(int(round(20 * self.p_go * (1 - self.p_reward_go)))) * 3  # code 3
            temp_comb1 = np.concatenate((temp_comb1, train_go_omission))
            if self.phase == 'cond':
                train_empty = np.ones(int(20 * self.p_empty* (1-self.p_reward_empty))) * 2  # code 2

                temp_comb1 = np.concatenate((temp_comb1, train_empty))
            elif self.phase == 'deg':
                train_unpredwater = np.ones(int(20 * self.p_empty* self.p_reward_empty)) * 4 # code 4
                temp_comb1 = np.concatenate((temp_comb1, train_unpredwater))
                train_empty = np.ones(int(20 * self.p_empty * (1 - self.p_reward_empty))) * 2  # code 2
                temp_comb1 = np.concatenate((temp_comb1, train_empty))
            elif self.phase == 'close':
                train_unpredwater = np.ones(int(20 * self.p_empty * self.p_reward_empty)) * 7  # code 4
                temp_comb1 = np.concatenate((temp_comb1, train_unpredwater))
                train_empty = np.ones(int(20 * self.p_empty * (1 - self.p_reward_empty))) * 2  # code 2
                temp_comb1 = np.concatenate((temp_comb1, train_empty))
            elif self.phase == 'far':
                train_unpredwater = np.ones(int(20 * self.p_empty * self.p_reward_empty)) * 8  # code 4
                temp_comb1 = np.concatenate((temp_comb1, train_unpredwater))
                train_empty = np.ones(int(20 * self.p_empty * (1 - self.p_reward_empty))) * 2  # code 2
                temp_comb1 = np.concatenate((temp_comb1, train_empty))
            elif self.phase == 'C_control':
                train_ccontrol = np.ones(int(20 * self.p_empty * self.p_reward_empty)) * 5  # code 4
                temp_comb1 = np.concatenate((temp_comb1, train_ccontrol))
                train_ccontrol_nogo = np.ones(int(20 * self.p_empty * (1 - self.p_reward_empty))) * 6  # code 2
                print(len(train_ccontrol_nogo))
                print(len(train_ccontrol))
                temp_comb1 = np.concatenate((temp_comb1, train_ccontrol_nogo))

            temp_comb2 = temp_comb1.copy()
            random.shuffle(temp_comb2)
            trialtypes[20*i:20*(i+1)] = temp_comb2
        self.trialstype = trialtypes
        print('================== Trial Types =================')
        print(trialtypes)
        for t in range(0, self.numtrials):
            print('================================================')
            print('trial number: ', t)

            d0 = self.ws0.cell(row=(self.ws0.max_row + 1), column=1, value=time.clock())
            d0 = self.ws0.cell(row=self.ws0.max_row, column=2, value=101) #trial start
            d0 = self.ws0.cell(row=self.ws0.max_row, column=3, value= 'trial{}'.format(int(trialtypes[t]))) #trial type

            d1 = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d1 = self.ws1.cell(row=self.ws1.max_row, column=2, value=101)  # trial start
            d1 = self.ws1.cell(row=self.ws1.max_row, column=3, value='trial{}'.format(int(trialtypes[t])))  # trial type
            # code: 0--go w rward; 1--no go; 2--empty; 3--go w/o reward; 4 --unpred water; 5 -- c control odor； 6 -- c control odor w/o
            if int(trialtypes[t]) in [7,8]:

                #           main training program
                self.run_trial_type(int(trialtypes[t]))


            else:
                self.check_licking_1spout(self.duration_rec_on_before)
                #           main training program
                self.run_trial_type(int(trialtypes[t]))

                self.check_licking_1spout(self.duration_rec_on_after)
            # self.settings.save_video.update_value(False):
            self.check_licking_1spout(self.duration_ITI[t])

            d0 = self.ws0.cell(row=(self.ws0.max_row + 1), column=1, value=time.clock())
            d0 = self.ws0.cell(row=self.ws0.max_row, column=2, value=100)  # end

            d1 = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d1 = self.ws1.cell(row=self.ws1.max_row, column=2, value=100) #end
            self.wb0.save(self.filename_0)
            self.wb1.save(self.filename_1)
        odors_cue.initiate()
        odors_cue.close()
        self.water_0.low()
        self.water_0.close()
        self.water_1.low()
        self.water_1.close()
        print('FINISHED ASSOCIATION TRAINING')

    def check_licking_1spout(self, interval,check_action=False):
        checkperiod = 0.002
        timeout = time.time() + interval
        reward_on = True # this is used for Pav or Operant
        right_lick_last = 0
        left_lick_last = 0
        count = 0
        while time.time() < timeout:
            right_lick = self.lick_0.read()
            # print('right lick', right_lick)

            left_lick = self.lick_1.read()
            # print('left lick', left_lick)
            if right_lick != right_lick_last:
                if right_lick:
                    print('Lick 0')
                    d0 = self.ws0.cell(row=(self.ws0.max_row + 1), column=1, value=time.clock())
                    d0= self.ws0.cell(row=self.ws0.max_row, column=2, value=11)

                # self.save_training()
                    if check_action:
                        count += 1
                else:
                    d0 = self.ws0.cell(row=(self.ws0.max_row + 1), column=1, value=time.clock())
                    d0 = self.ws0.cell(row=self.ws0.max_row, column=2, value=10)

            else:
                pass

            if left_lick != left_lick_last:
                if left_lick:
                    print('Lick 1')
                    d1 = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
                    d1 = self.ws1.cell(row=self.ws1.max_row, column=2, value=11)

                    # self.save_training()
                    if check_action:
                        count += 1
                else:
                    d1 = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
                    d1 = self.ws1.cell(row=self.ws1.max_row, column=2, value=10)

            else:
                pass
            right_lick_last = right_lick
            left_lick_last = left_lick
            time.sleep(checkperiod)
        if check_action and count >= self.licknum:

            print('licking activate reward')
        elif check_action and count < self.licknum:
            print('not enough licking')
            reward_on = False
        return reward_on

    def run_trial_type(self, types):
        odor_on = False
        reward_on = False
        is_control = False
        if types == 0:
            print('go trial ' + str(int(self.counter[types])))
            odor_on = True # contingency odor comes
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action)
            reward_on = self.check_licking_1spout(self.duration_action_window, self.operant)  ### can be a problem
            self.run_reward_module(reward_on, w_code)

        elif types == 1:
            print('no go trial ' + str(int(self.counter[types])))

            odor_on = True
            is_go = False
            r_code = [141, 140]
            w_code = [61, 60]
            self.run_odor_module(odor_on,is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action+self.duration_action_window)
            reward_on = False
            self.run_reward_module(reward_on, w_code)
        elif types == 5:
            print('control C trial ' + str(int(self.counter[types])))

            odor_on = True
            is_go = False
            is_control = True
            r_code = [161, 160]
            w_code = [51, 50]
            self.run_odor_module(odor_on,is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action+self.duration_action_window)
            reward_on = self.check_licking_1spout(self.duration_action_window, self.operant)
            self.run_reward_module(reward_on, w_code)

        elif types == 2:
            print('empty trial ' + str(int(self.counter[types])))
            # odor false: control odor comes
            w_code = [71, 70]# not used
           # self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_on)
            self.check_licking_1spout(self.duration_odor_to_action+self.duration_action_window)
            reward_on = False
            self.run_reward_module(reward_on, w_code)

        elif types == 3:
            print('go omission trial ' + str(int(self.counter[types])))
            odor_on = True # contingency odor comes
            is_go = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action + self.duration_action_window)
            self.run_reward_module(reward_on, w_code)

        elif types == 6:
            print('c odor omission ' + str(int(self.counter[types])))
            odor_on = True # contingency odor comes
            is_go = False
            is_control = True
            r_code = [161, 160]
            w_code = [51, 50]
            self.run_odor_module(odor_on, is_go, is_control, r_code)
            self.check_licking_1spout(self.duration_odor_to_action + self.duration_action_window)
            self.run_reward_module(reward_on, w_code)

        elif types == 4:
            print('unpredicted water trial ' + str(int(self.counter[types])))
            w_code = [51, 50]
            self.check_licking_1spout(self.duration_odor_on)
            self.check_licking_1spout(self.duration_odor_to_action)
            reward_on = self.check_licking_1spout(self.duration_action_window, self.operant)
            self.run_reward_module(reward_on, w_code)

        elif types == 7:
            print('close unpredicted water trial ' + str(int(self.counter[types])))
            w_code = [51, 50]
            reward_on = self.check_licking_1spout(0, self.operant)
            self.run_reward_module(reward_on, w_code)
            self.check_licking_1spout(self.duration_odor_on+2)
            self.check_licking_1spout(self.duration_action_window)

        elif types == 8:
            print('far unpredicted water trial ' + str(int(self.counter[types])))
            w_code = [51, 50]
            self.check_licking_1spout(self.duration_odor_on)
            self.check_licking_1spout(self.duration_odor_to_action+2)
            reward_on = self.check_licking_1spout(self.duration_action_window, self.operant)
            self.run_reward_module(reward_on, w_code)
        self.counter[types] += 1



    def run_odor_module(self, odor_on, is_go, is_control, r_code):
        if odor_on:
            print('opening odor port')
            if is_go and not is_control:
                self.reward_odor.high()
            elif not is_go and not is_control:
                self.non_reward_odor.high()
            elif is_control:
                self.control_odor.high()
            # self.OdorOnCopy.high()  # ？？？
            d0 = self.ws0.cell(row=(self.ws0.max_row + 1), column=1, value=time.clock())
            d0 = self.ws0.cell(row=self.ws0.max_row, column=2, value=r_code[0])
            d1 = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d1 = self.ws1.cell(row=self.ws1.max_row, column=2, value=r_code[0])
            # self.save_training()
            self.check_licking_1spout(self.duration_odor_on)
            print('closing odor port')
            if is_go and not is_control:
                self.reward_odor.low()
            elif not is_go and not is_control:
                self.non_reward_odor.low()
            elif is_control:
                self.control_odor.low()
            d0 = self.ws0.cell(row=(self.ws0.max_row + 1), column=1, value=time.clock())
            d0 = self.ws0.cell(row=self.ws0.max_row, column=2, value=r_code[1])
            d1 = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d1 = self.ws1.cell(row=self.ws1.max_row, column=2, value=r_code[1])
        else:
            self.check_licking_1spout(self.duration_odor_on)

    def run_reward_module(self, reward_on, w_code):
        if reward_on:
            # modify! give water if licks three times within 1 s
            print('opening water valve')
            self.water_0.high()
            self.water_1.high()

            d0 = self.ws0.cell(row=(self.ws0.max_row + 1), column=1, value=time.clock())
            d0 = self.ws0.cell(row=self.ws0.max_row, column=2, value=w_code[0])
            d1 = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d1 = self.ws1.cell(row=self.ws1.max_row, column=2, value=w_code[0])
            self.check_licking_1spout(self.duration_water_large)  # this parameter hasn't een defined
            print('closing water valve')
            self.water_0.low()
            self.water_1.low()
            d0 = self.ws0.cell(row=(self.ws0.max_row + 1), column=1, value=time.clock())
            d0 = self.ws0.cell(row=self.ws0.max_row, column=2, value=w_code[1])
            d1 = self.ws1.cell(row=(self.ws1.max_row + 1), column=1, value=time.clock())
            d1 = self.ws1.cell(row=self.ws1.max_row, column=2, value=w_code[1])
        else:
            self.check_licking_1spout(self.duration_water_large)


#     def camera_action(self):
#         '''
#         format the image properly
#         '''
#         try:
#             wide_image = self.wide_cam.read()
#             wide_image_data = self.wide_cam.to_numpy(wide_image)
#             self.wide_disp_queue.put(wide_image_data)
#
#             if self.settings.save_video.value() and self.settings.in_trial.value() :
#                 self.recorder.save_frame(self.settings.filename.value(),wide_image)
#
#         except Exception as ex:
#             print('Error : %s' % ex)





test = SelinaTraining()
print('start')
test.run()


